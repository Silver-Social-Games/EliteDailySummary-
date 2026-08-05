"""
Apply Top 50 player swaps and regenerate management brief.

Swaps 10 weakest (by NP 30d) for the recommended unlocked replacements.
"""
from __future__ import annotations

import csv
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from daily_summary.generate_daily_elite_summary import PROJECT_ID, get_client, run_query

MODULE_DIR = Path(__file__).resolve().parent
DATA_DIR = MODULE_DIR / "data"
EXPORT_DIR = MODULE_DIR / "exports"
SOURCE = DATA_DIR / "vip-event-top50-players.xlsx"
SHEET = "VIP Event - Top 50 Players "
HDR_ROW = 8
DATA_START = 9
DATA_END = 58

SWAPS: dict[int, int] = {
    # out_aid: in_aid
    368254306: 383173826,   # Sally Davis -> Cassie Cannon
    384841200: 204544406,   # Yarkmealer Dozier -> Cathleen Gendreau
    212686132: 202236759,   # Maria Sieger -> Rosa Espinoza
    243814714: 53372335,    # Jesus Canez -> Maurice Davis
    359250867: 351574165,   # Jason Slater -> Joe Julian Martinez
    382659497: 401426364,   # Allison Stansberry -> Jason Galloway
    358663705: 432965821,   # Aleida Smith -> Richard Nichols
    230578311: 220994976,   # Brandon Reed -> Quorlisha Benifield
    216725336: 189473165,   # Belinda Stewart -> Kimberley Rego
    381182250: 147754388,   # Thuy Nguyen -> Tamikia Rosa
    134550509: 271692626,   # Susan Burton -> Brent Lindstrom (locked/self-exclusion)
}


def parse_money(raw) -> float:
    if raw is None:
        return 0.0
    s = str(raw).strip().replace("$", "").replace(",", "").replace('"', "")
    if not s:
        return 0.0
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    try:
        v = float(s)
    except ValueError:
        return 0.0
    return -v if neg else v


def parse_hold(raw) -> float | None:
    if raw is None:
        return None
    s = str(raw).strip().replace("%", "")
    if not s:
        return None
    try:
        v = float(s)
    except ValueError:
        return None
    return v / 100 if v > 1 else v


def load_csv_players() -> dict[int, dict]:
    path = DATA_DIR / "players-table-latest.csv"
    out: dict[int, dict] = {}
    with path.open(encoding="utf-16", newline="") as f:
        for row in csv.DictReader(f, delimiter="\t"):
            aid = int(row["account_id"])
            out[aid] = {
                "name": f"{row.get('first_name', '').strip()} {row.get('last_name', '').strip()}".strip(),
                "age": int(row["Age"]) if str(row.get("Age", "")).isdigit() else row.get("Age"),
                "state": row.get("Reg. State", ""),
                "agent": row.get("agent_name (group)", ""),
                "net_lt": parse_money(row.get("Max Value Net Purchase")),
                "hold": parse_hold(row.get("Hold %")),
            }
    return out


def fetch_bq_metrics(aids: list[int], report_date: date) -> dict[int, dict]:
    if not aids:
        return {}
    rd = report_date.isoformat()
    d30 = (report_date - timedelta(days=29)).isoformat()
    d60 = (report_date - timedelta(days=59)).isoformat()
    in_list = ", ".join(str(a) for a in aids)
    tbl = f"{PROJECT_ID}.jackpota_agg.daily_player_revenue_kpis"
    sql = f"""
WITH daily AS (
  SELECT account_id AS aid, date,
    SUM(CAST(purchased AS FLOAT64)) AS purchased,
    SUM(CAST(purchased AS FLOAT64)-CAST(redeemed AS FLOAT64)
      -CAST(chargeback AS FLOAT64)-CAST(refunds AS FLOAT64)) AS np
  FROM `{tbl}` WHERE account_id IN ({in_list})
  GROUP BY account_id, date
)
SELECT aid,
  ROUND(SUM(np), 2) AS net_lt,
  ROUND(SUM(IF(date BETWEEN DATE '{d30}' AND DATE '{rd}', np, 0)), 2) AS net_30d,
  ROUND(SUM(IF(date BETWEEN DATE '{d60}' AND DATE '{rd}', np, 0)), 2) AS net_60d,
  ROUND(SAFE_DIVIDE(SUM(np), NULLIF(SUM(purchased), 0)), 4) AS hold
FROM daily GROUP BY aid
"""
    return {int(r["aid"]): r for r in run_query(get_client(), sql)}


def apply_swaps(source: Path) -> list[tuple[int, int, str, str]]:
    csv_players = load_csv_players()
    rd = date.today() - timedelta(days=1)
    in_aids = list(SWAPS.values())
    bq = fetch_bq_metrics(in_aids, rd)

    wb = load_workbook(source)
    ws = wb[SHEET]
    log: list[tuple[int, int, str, str]] = []

    for row in range(DATA_START, DATA_END + 1):
        aid = ws.cell(row, 2).value
        if aid is None:
            continue
        aid = int(aid)
        if aid not in SWAPS:
            continue
        new_aid = SWAPS[aid]
        old_name = ws.cell(row, 3).value
        p = csv_players.get(new_aid)
        if not p:
            raise ValueError(f"AID {new_aid} not found in players CSV")
        b = bq.get(new_aid, {})

        ws.cell(row, 2).value = new_aid
        ws.cell(row, 3).value = p["name"]
        ws.cell(row, 4).value = p["age"]
        ws.cell(row, 5).value = p["state"]
        ws.cell(row, 6).value = p["agent"]
        ws.cell(row, 7).value = float(b.get("net_lt") or p["net_lt"])
        if ws.max_column >= 8 and ws.cell(HDR_ROW, 8).value and "30" in str(ws.cell(HDR_ROW, 8).value):
            ws.cell(row, 8).value = float(b.get("net_30d") or 0)
            ws.cell(row, 9).value = float(b.get("net_60d") or 0)
            hold_col = 10
        else:
            hold_col = 8 if "hold" in str(ws.cell(HDR_ROW, 8).value or "").lower() else 10
        hold = b.get("hold")
        if hold is None:
            hold = p["hold"]
        ws.cell(row, hold_col).value = float(hold) if hold is not None else None

        log.append((aid, new_aid, str(old_name), p["name"]))

    wb.save(source)
    return log


def main() -> None:
    if not SOURCE.exists():
        raise FileNotFoundError(f"Source not found: {SOURCE}")

    log = apply_swaps(SOURCE)
    print(f"Updated {SOURCE}\n")
    for old_aid, new_aid, old_name, new_name in log:
        print(f"  OUT {old_aid} {old_name}")
        print(f"  IN  {new_aid} {new_name}\n")

    # Regenerate management brief
    from generate_top50_management_brief import generate

    out = EXPORT_DIR / "VIP Event - Top 50 - Management Brief.xlsx"
    path, _ = generate(SOURCE, date.today() - timedelta(days=1), out)
    print(f"Regenerated: {path}")

    try:
        import os
        os.startfile(path)
    except OSError:
        pass


if __name__ == "__main__":
    main()
