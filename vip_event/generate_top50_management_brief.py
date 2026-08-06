"""
Generate management-ready VIP Event Top 50 Excel brief (new design, built from scratch).

Usage:
  python vip_event/generate_top50_management_brief.py
  python vip_event/generate_top50_management_brief.py --date YYYY-MM-DD
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from elite_lib import PROJECT_ID, get_client, run_query

MODULE_DIR = Path(__file__).resolve().parent
DATA_DIR = MODULE_DIR / "data"
EXPORT_DIR = MODULE_DIR / "exports"

DEFAULT_SOURCE = Path(
    r"c:\Users\Owner\OneDrive - Silver Social Games\Desktop\VIP\VIP Event"
    r"\VIP Event - Top 50 Players .xlsx"
)
SHEET_NAMES = ("VIP Event - Top 50 Players ", "Top 50 Brief")

# Layout
HDR_ROW = 9
DATA_START = 10
DATA_END = 59
TOTAL_ROW = 60
LAST_COL = 15

# Roster column indices
COL_RANK, COL_AID, COL_NAME = 1, 2, 3
COL_AGE, COL_STATE, COL_AGENT = 4, 5, 6
COL_REG, COL_SEN_DAYS, COL_SEN_GROUP = 7, 8, 9
COL_NP_LT, COL_NP_30, COL_NP_60 = 10, 11, 12
COL_PURCH_30, COL_HOLD, COL_STATUS = 13, 14, 15

DATE_FMT = "dd-mmm-yyyy"
WARN_FILL_COLOR = "FFFFF8E8"

# Seniority buckets — parity with Daily_Agg `DaysFromFTP` / `Seniority`
SENIORITY_BUCKET_ORDER = (
    "First Day",
    "First 2-7d",
    "First 8-30d",
    "First 31-90d",
    "First 91-180",
    "First 181-365",
    "+1Y",
)

# Executive palette
NAVY = "FF1B2A4A"
NAVY_MID = "FF2C3E6E"
ACCENT = "FF3D6B9E"
MUTED = "FF6B7C93"
INK = "FF1A1A2E"
WHITE = "FFFFFFFF"
SURFACE = "FFF7F9FC"
SURFACE_ALT = "FFF0F4F8"
HEADER_BG = "FFE8EEF5"
RULE = "FFE2E8F0"
TOTAL_BG = "FFD4E0ED"
ORANGE = "FFED7D31"          # metric highlight (Excel accent orange)
ORANGE_LIGHT = "FFFFF4E8"    # KPI card background
ORANGE_DARK = "FFC65911"     # section accent bar

SIDE_THIN = Side(style="thin", color=RULE)
BORDER = Border(left=SIDE_THIN, right=SIDE_THIN, top=SIDE_THIN, bottom=SIDE_THIN)
BORDER_BOTTOM_ACCENT = Border(bottom=Side(style="medium", color=ACCENT))

MONEY_FMT = '"$"#,##0'
PCT_FMT = "0.0%"

FONT_H1 = Font(name="Calibri", size=26, bold=True, color=WHITE)
FONT_H2 = Font(name="Calibri", size=12, color="FFB8C9E0")
FONT_META = Font(name="Calibri", size=10, color=MUTED[2:])
FONT_KPI_VAL = Font(name="Calibri", size=22, bold=True, color=ORANGE[2:])
FONT_KPI_LBL = Font(name="Calibri", size=9, bold=True, color=MUTED[2:])
FONT_KPI_VAL_INK = Font(name="Calibri", size=22, bold=True, color=INK[2:])
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color=NAVY[2:])
FONT_TH = Font(name="Calibri", size=10, bold=True, color=NAVY[2:])
FONT_BODY = Font(name="Calibri", size=11, color=INK[2:])
FONT_NAME = Font(name="Calibri", size=11, bold=True, color=INK[2:])
FONT_RANK = Font(name="Calibri", size=11, bold=True, color=ACCENT[2:])
FONT_MONEY = Font(name="Calibri", size=11, bold=True, color=INK[2:])
FONT_TOTAL = Font(name="Calibri", size=11, bold=True, color=NAVY[2:])
FONT_STATE = Font(name="Calibri", size=13, bold=True, color=INK[2:])

# Presentation slides (16:9 landscape, one sheet = one slide)
SLIDE_COLS = 16
SLIDE_ROSTER_ROWS = 25  # per column block on slide 2


def agent_breakdown(players: list[Player]) -> list[tuple[str, int, float]]:
    counts = Counter(p.agent for p in players if p.agent)
    total = sum(counts.values()) or 1
    return [(name, n, n / total) for name, n in counts.most_common(6)]


def _configure_slide_page(ws, last_row: int, last_col: int = SLIDE_COLS) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"
    for col in range(1, last_col + 1):
        if get_column_letter(col) not in ws.column_dimensions:
            ws.column_dimensions[get_column_letter(col)].width = 10


def _slide_hero(
    ws,
    subtitle: str,
    report_date: date,
    active_count: int,
    slide_num: int,
    last_col: int = SLIDE_COLS,
) -> int:
    """Navy header band; returns first content row below accent."""
    navy = PatternFill("solid", fgColor=NAVY)
    navy_mid = PatternFill("solid", fgColor=NAVY_MID)
    orange = PatternFill("solid", fgColor=ORANGE)
    white = PatternFill("solid", fgColor=WHITE)

    f_title = Font(name="Calibri", size=32, bold=True, color=WHITE)
    f_sub = Font(name="Calibri", size=15, color="FFB8C9E0")
    f_meta = Font(name="Calibri", size=10, color=MUTED[2:])
    f_slide = Font(name="Calibri", size=10, bold=True, color="FFB8C9E0")

    as_of = report_date.strftime("%b %d, %Y")
    merge_style(ws, 1, 1, 2, last_col - 1, "VIP Event", f_title, navy,
                Alignment(horizontal="left", vertical="center", indent=1))
    set_cell(ws, 1, last_col, f"{slide_num} / 2", f_slide, navy,
             Alignment(horizontal="right", vertical="center"))
    set_cell(ws, 2, last_col, "", f_slide, navy)
    merge_style(ws, 3, 1, 3, last_col, subtitle, f_sub, navy,
                Alignment(horizontal="left", vertical="center", indent=1))
    merge_style(
        ws, 4, 1, 4, last_col,
        f"As of {as_of}   ·   {active_count} active Elite invitees   ·   locked / self-exclusion excluded",
        f_meta, navy_mid, Alignment(horizontal="left", vertical="center", indent=1),
    )
    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 5
    for c in range(1, last_col + 1):
        set_cell(ws, 5, c, fill=orange)
    ws.row_dimensions[6].height = 10
    return 7


def _slide_kpi_row(ws, row: int, cards: list[tuple[int, int, str, str]], last_col: int = SLIDE_COLS) -> int:
    """Write merged KPI cards; returns row after labels."""
    orange_light = PatternFill("solid", fgColor=ORANGE_LIGHT)
    white = PatternFill("solid", fgColor=WHITE)
    f_val = Font(name="Calibri", size=24, bold=True, color=ORANGE[2:])
    f_lbl = Font(name="Calibri", size=8, bold=True, color=MUTED[2:])
    card_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lbl_align = Alignment(horizontal="center", vertical="top", wrap_text=True)

    for c1, c2, val, lbl in cards:
        merge_style(ws, row, c1, row, c2, val, f_val, orange_light, card_align)
        merge_style(ws, row + 1, c1, row + 1, c2, lbl, f_lbl, white, lbl_align)
        for c in range(c1, c2 + 1):
            set_cell(ws, row, c, border=BORDER)
            set_cell(ws, row + 1, c, border=BORDER)

    ws.row_dimensions[row].height = 40
    ws.row_dimensions[row + 1].height = 28
    return row + 3


def _slide_panel_table(
    ws,
    start_row: int,
    c1: int,
    c2: int,
    title: str,
    headers: list[tuple[int, str]],
    rows: list[tuple],
    col_fmts: dict[int, str | None] | None = None,
) -> int:
    """Compact panel with navy title bar; returns row after table."""
    navy = PatternFill("solid", fgColor=NAVY)
    hdr_fill = PatternFill("solid", fgColor=HEADER_BG)
    white = PatternFill("solid", fgColor=WHITE)
    alt_fill = PatternFill("solid", fgColor=SURFACE_ALT)
    f_panel = Font(name="Calibri", size=11, bold=True, color=WHITE)
    f_th = Font(name="Calibri", size=9, bold=True, color=NAVY[2:])
    f_body = Font(name="Calibri", size=10, color=INK[2:])

    merge_style(ws, start_row, c1, start_row, c2, f"  {title}", f_panel, navy,
                Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[start_row].height = 22

    hr = start_row + 1
    for col, label in headers:
        set_cell(ws, hr, col, label, f_th, hdr_fill,
                 Alignment(horizontal="center" if col > headers[0][0] else "left", vertical="center"),
                 border=BORDER)
    ws.row_dimensions[hr].height = 20

    fmts = col_fmts or {}
    for i, row_data in enumerate(rows):
        r = hr + 1 + i
        fill = alt_fill if i % 2 else white
        for j, (col, val) in enumerate(row_data):
            halign = "center" if j > 0 else "left"
            set_cell(ws, r, col, val, f_body, fill, Alignment(horizontal=halign, vertical="center"),
                     fmts.get(col), BORDER)
        ws.row_dimensions[r].height = 18
    return hr + 1 + len(rows)


def build_slide1_cohort(
    wb: Workbook,
    players: list[Player],
    summary: dict,
    report_date: date,
    active_count: int,
) -> None:
    """Slide 1 — cohort story: value, tenure, geography, agent coverage."""
    ws = wb.create_sheet("Slide 1 — Cohort", 0)
    lc = SLIDE_COLS
    content = _slide_hero(ws, "Top 50 Elite Invitees  ·  Cohort Overview", report_date, active_count, 1, lc)

    med_days = summary.get("median_seniority_days")
    med_group = summary.get("median_seniority_group", "—")
    med_tenure = f"{med_days:,} d" if med_days is not None else "—"

    content = _slide_kpi_row(ws, content, [
        (1, 4, str(active_count), "ACTIVE INVITEES"),
        (5, 8, f"${summary['sum_np_30d']:,.0f}", "COHORT NET PURCHASE · 30 DAYS"),
        (9, 12, f"${summary['sum_np_lt']:,.0f}", "COHORT NET PURCHASE · LIFETIME"),
        (13, lc, med_tenure, f"MEDIAN TENURE · {med_group}"),
    ], lc)

    content = _slide_kpi_row(ws, content, [
        (1, 5, f"${summary['sum_purchased_30d']:,.0f}", "PURCHASED · 30 DAYS"),
        (6, 10, f"${summary['sum_np_60d']:,.0f}", "NET PURCHASE · 60 DAYS"),
        (11, lc, f"{summary['avg_hold']:.1%}", "AVG HOLD % · LIFETIME"),
    ], lc)

    sen_rows = [
        [(1, name), (2, n), (3, pct)]
        for name, n, pct in summary.get("seniority", seniority_breakdown(players))
    ]
    state_rows = [
        [(9, name), (10, n), (11, pct)]
        for name, n, pct in summary.get("states", state_breakdown(players))[:8]
    ]
    max_len = max(len(sen_rows), len(state_rows), 1)
    panel_row = content
    navy = PatternFill("solid", fgColor=NAVY)
    hdr_fill = PatternFill("solid", fgColor=HEADER_BG)
    white = PatternFill("solid", fgColor=WHITE)
    alt_fill = PatternFill("solid", fgColor=SURFACE_ALT)
    f_panel = Font(name="Calibri", size=11, bold=True, color=WHITE)
    f_th = Font(name="Calibri", size=9, bold=True, color=NAVY[2:])
    f_body = Font(name="Calibri", size=10, color=INK[2:])

    merge_style(ws, panel_row, 1, panel_row, 7, "  TENURE · DAYS FROM FTP", f_panel, navy,
                Alignment(horizontal="left", vertical="center", indent=1))
    merge_style(ws, panel_row, 9, panel_row, lc, "  GEOGRAPHY · STATE MIX", f_panel, navy,
                Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[panel_row].height = 22
    hr = panel_row + 1
    for col, label in ((1, "Cohort"), (2, "Players"), (3, "Share")):
        set_cell(ws, hr, col, label, f_th, hdr_fill, Alignment(horizontal="center" if col > 1 else "left"), border=BORDER)
    for col, label in ((9, "State"), (10, "Players"), (11, "Share")):
        set_cell(ws, hr, col, label, f_th, hdr_fill, Alignment(horizontal="center" if col > 1 else "left"), border=BORDER)
    ws.row_dimensions[hr].height = 20

    for i in range(max_len):
        r = hr + 1 + i
        fill = alt_fill if i % 2 else white
        if i < len(sen_rows):
            name, n, pct = sen_rows[i][0][1], sen_rows[i][1][1], sen_rows[i][2][1]
            set_cell(ws, r, 1, name, f_body, fill, Alignment(horizontal="left"), border=BORDER)
            set_cell(ws, r, 2, n, f_body, fill, Alignment(horizontal="center"), "0", BORDER)
            set_cell(ws, r, 3, pct, f_body, fill, Alignment(horizontal="center"), PCT_FMT, BORDER)
        if i < len(state_rows):
            name, n, pct = state_rows[i][0][1], state_rows[i][1][1], state_rows[i][2][1]
            set_cell(ws, r, 9, name, f_body, fill, Alignment(horizontal="left"), border=BORDER)
            set_cell(ws, r, 10, n, f_body, fill, Alignment(horizontal="center"), "0", BORDER)
            set_cell(ws, r, 11, pct, f_body, fill, Alignment(horizontal="center"), PCT_FMT, BORDER)
        ws.row_dimensions[r].height = 18

    agent_row = hr + max_len + 2
    agents = agent_breakdown(players)
    agent_data = [[(1, name), (2, n), (3, pct)] for name, n, pct in agents]
    last_row = _slide_panel_table(
        ws, agent_row, 1, 7,
        "AGENT COVERAGE",
        [(1, "Agent"), (2, "Players"), (3, "Share")],
        agent_data,
        {3: PCT_FMT},
    )

    takeaway_row = last_row + 1
    top = summary.get("top_state", "—")
    merge_style(
        ws, takeaway_row, 1, takeaway_row, lc,
        (
            f"Managed Elite book  ·  Median age {summary['median_age']:.0f}  ·  "
            f"Largest state: {top} ({summary['top_state_n']} players)  ·  "
            "Tenure groups match warehouse DaysFromFTP buckets"
        ),
        Font(name="Calibri", size=9, italic=True, color=MUTED[2:]),
        PatternFill("solid", fgColor=WHITE),
        Alignment(horizontal="left", indent=1),
    )
    ws.row_dimensions[takeaway_row].height = 18

    widths = {1: 18, 2: 9, 3: 9, 4: 2, 5: 2, 6: 2, 7: 2, 8: 2,
              9: 16, 10: 9, 11: 9, 12: 2, 13: 2, 14: 2, 15: 2, 16: 2}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    _configure_slide_page(ws, takeaway_row, lc)


def _write_roster_block(
    ws,
    players: list[Player],
    start_row: int,
    cols: dict[str, int],
) -> int:
    """Write one 25-player roster block; cols maps keys to column indices."""
    c_rank = cols["rank"]
    c_name = cols["name"]
    c_state = cols["state"]
    c_np30 = cols["np30"]
    c_days = cols["days"]
    c_end = max(cols.values())

    hdr_fill = PatternFill("solid", fgColor=HEADER_BG)
    navy = PatternFill("solid", fgColor=NAVY)
    white = PatternFill("solid", fgColor=WHITE)
    alt_fill = PatternFill("solid", fgColor=SURFACE_ALT)
    f_th = Font(name="Calibri", size=9, bold=True, color=NAVY[2:])
    f_rank = Font(name="Calibri", size=9, bold=True, color=ACCENT[2:])
    f_name = Font(name="Calibri", size=9, bold=True, color=INK[2:])
    f_body = Font(name="Calibri", size=9, color=INK[2:])
    f_money = Font(name="Calibri", size=9, bold=True, color=INK[2:])

    for col, label in (
        (c_rank, "#"), (c_name, "Player"), (c_state, "State"),
        (c_np30, "NP 30d"), (c_days, "Days FTP"),
    ):
        set_cell(ws, start_row, col, label, f_th, hdr_fill,
                 Alignment(horizontal="center" if col != c_name else "left", vertical="center"),
                 border=BORDER)
    ws.row_dimensions[start_row].height = 18

    for i, p in enumerate(players):
        r = start_row + 1 + i
        fill = alt_fill if i % 2 else white
        set_cell(ws, r, c_rank, p.rank, f_rank, fill, Alignment(horizontal="center"), border=BORDER)
        set_cell(ws, r, c_name, p.name, f_name, fill, Alignment(horizontal="left"), border=BORDER)
        set_cell(ws, r, c_state, p.state, f_body, fill, Alignment(horizontal="left"), border=BORDER)
        set_cell(ws, r, c_np30, p.net_30d, f_money, fill, Alignment(horizontal="right"), MONEY_FMT, BORDER)
        days_val = p.seniority_days if p.seniority_days is not None else "—"
        days_fmt = "#,##0" if isinstance(days_val, int) else None
        set_cell(ws, r, c_days, days_val, f_body, fill, Alignment(horizontal="center"), days_fmt, BORDER)
        ws.row_dimensions[r].height = 15

    return start_row + len(players)


def build_slide2_roster(
    wb: Workbook,
    players: list[Player],
    summary: dict,
    report_date: date,
    active_count: int,
) -> None:
    """Slide 2 — full invitee list, dual-column for readability."""
    ws = wb.create_sheet("Slide 2 — Roster", 1)
    lc = SLIDE_COLS
    table_row = _slide_hero(ws, "Top 50 Elite Invitees  ·  Player Roster", report_date, active_count, 2, lc)

    table_row = _slide_kpi_row(ws, table_row, [
        (1, 4, f"${summary['sum_np_30d']:,.0f}", "TOTAL NP · 30 DAYS"),
        (5, 8, f"${summary['sum_np_lt']:,.0f}", "TOTAL NP · LIFETIME"),
        (9, 12, f"${summary['sum_purchased_30d']:,.0f}", "PURCHASED · 30 DAYS"),
        (13, lc, f"{summary['avg_hold']:.1%}", "AVG HOLD %"),
    ], lc)

    navy = PatternFill("solid", fgColor=NAVY)
    f_panel = Font(name="Calibri", size=11, bold=True, color=WHITE)
    merge_style(ws, table_row, 1, table_row, 7, "  RANKS 1 – 25", f_panel, navy,
                Alignment(horizontal="left", vertical="center", indent=1))
    merge_style(ws, table_row, 9, table_row, lc, "  RANKS 26 – 50", f_panel, navy,
                Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[table_row].height = 22

    hdr_row = table_row + 1
    left_cols = {"rank": 1, "name": 2, "state": 3, "np30": 4, "days": 5}
    right_cols = {"rank": 9, "name": 10, "state": 11, "np30": 12, "days": 13}
    half = len(players) // 2
    last_left = _write_roster_block(ws, players[:half], hdr_row, left_cols)
    last_right = _write_roster_block(ws, players[half:], hdr_row, right_cols)
    last_data = max(last_left, last_right)

    foot = last_data + 2
    merge_style(
        ws, foot, 1, foot, lc,
        "Net purchase = purchased − redeemed − chargebacks − refunds   ·   Full detail on “Detail — Full Data” tab",
        Font(name="Calibri", size=9, italic=True, color=MUTED[2:]),
        PatternFill("solid", fgColor=WHITE),
        Alignment(horizontal="left", indent=1),
    )
    ws.row_dimensions[foot].height = 16

    widths = {1: 4, 2: 17, 3: 11, 4: 11, 5: 8, 6: 2, 7: 2, 8: 2,
              9: 4, 10: 17, 11: 11, 12: 11, 13: 8, 14: 2, 15: 2, 16: 2}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    _configure_slide_page(ws, foot, lc)


def build_presentation_slides(
    wb: Workbook,
    players: list[Player],
    summary: dict,
    report_date: date,
    active_count: int,
) -> None:
    build_slide1_cohort(wb, players, summary, report_date, active_count)
    build_slide2_roster(wb, players, summary, report_date, active_count)


@dataclass
class Player:
    rank: int
    aid: int
    name: str
    age: int | float | None
    state: str
    agent: str
    net_lt: float
    hold: float | None
    net_30d: float = 0.0
    net_60d: float = 0.0
    reg_date: date | None = None
    ftp_date: date | None = None
    seniority_days: int | None = None
    seniority_group: str = "—"
    purchased_30d: float = 0.0
    invite_status: str = "Active"


def _parse_bq_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, date):
        return val
    return date.fromisoformat(str(val)[:10])


def seniority_group(days: int | None) -> str:
    """Match `Daily_Agg_Per_Player_Query_v1.sql` Seniority CASE on DaysFromFTP."""
    if days is None or days < 0:
        return "—"
    if days == 0:
        return "First Day"
    if days <= 6:
        return "First 2-7d"
    if days < 30:
        return "First 8-30d"
    if days < 90:
        return "First 31-90d"
    if days < 180:
        return "First 91-180"
    if days < 365:
        return "First 181-365"
    return "+1Y"


def format_seniority_days(days: int | None) -> str:
    if days is None or days < 0:
        return "—"
    return f"{days:,} d"


def median_seniority_days(players: list[Player]) -> int | None:
    days = [p.seniority_days for p in players if p.seniority_days is not None]
    if not days:
        return None
    days.sort()
    return days[len(days) // 2]


def invite_status_label(locked, lock_reason: str | None) -> str:
    if locked is True or str(locked).lower() in ("true", "1"):
        reason = (lock_reason or "").lower()
        if "exclusion" in reason:
            return "Self-exclusion"
        return "Locked"
    return "Active"


def build_enrich_sql(aids: list[int], report_date: date) -> str:
    rd = report_date.isoformat()
    d30 = (report_date - timedelta(days=29)).isoformat()
    d60 = (report_date - timedelta(days=59)).isoformat()
    in_list = ", ".join(str(a) for a in aids)
    kpi_tbl = f"{PROJECT_ID}.jackpota_agg.daily_player_revenue_kpis"
    ps_tbl = f"{PROJECT_ID}.dbt_marketing_mart.player_stats_daily"
    acct_tbl = f"{PROJECT_ID}.transactional_data.uam_accounts"
    return f"""
WITH daily AS (
  SELECT
    account_id AS aid,
    date,
    SUM(CAST(purchased AS FLOAT64)) AS purchased,
    SUM(
      CAST(purchased AS FLOAT64) - CAST(redeemed AS FLOAT64)
      - CAST(chargeback AS FLOAT64) - CAST(refunds AS FLOAT64)
    ) AS net_purchases
  FROM `{kpi_tbl}`
  WHERE account_id IN ({in_list})
  GROUP BY account_id, date
),
kpi AS (
  SELECT
    aid,
    ROUND(SUM(net_purchases), 2) AS net_lt,
    ROUND(SUM(IF(date BETWEEN DATE '{d30}' AND DATE '{rd}', net_purchases, 0)), 2) AS net_30d,
    ROUND(SUM(IF(date BETWEEN DATE '{d60}' AND DATE '{rd}', net_purchases, 0)), 2) AS net_60d,
    ROUND(SUM(IF(date BETWEEN DATE '{d30}' AND DATE '{rd}', purchased, 0)), 2) AS purchased_30d,
    ROUND(SAFE_DIVIDE(SUM(net_purchases), NULLIF(SUM(purchased), 0)), 4) AS hold_lt
  FROM daily
  GROUP BY aid
),
ftp_kpi AS (
  SELECT aid, MIN(date) AS ftp_date_kpi
  FROM daily
  WHERE purchased > 0
  GROUP BY aid
),
profile AS (
  SELECT
    account_id AS aid,
    MAX(DATE(reg_date)) AS reg_date,
    MAX(DATE(ftp_date)) AS ftp_date_ps
  FROM `{ps_tbl}`
  WHERE account_id IN ({in_list})
  GROUP BY account_id
),
accounts AS (
  SELECT id AS aid, locked, lock_reason
  FROM `{acct_tbl}`
  WHERE id IN ({in_list})
)
SELECT
  k.aid,
  k.net_lt,
  k.net_30d,
  k.net_60d,
  k.purchased_30d,
  k.hold_lt,
  p.reg_date,
  COALESCE(p.ftp_date_ps, fk.ftp_date_kpi) AS ftp_date,
  a.locked,
  a.lock_reason
FROM kpi k
LEFT JOIN profile p ON k.aid = p.aid
LEFT JOIN ftp_kpi fk ON k.aid = fk.aid
LEFT JOIN accounts a ON k.aid = a.aid
"""


def fetch_enrich_metrics(aids: list[int], report_date: date) -> dict[int, dict]:
    if not aids:
        return {}
    rows = run_query(get_client(), build_enrich_sql(aids, report_date))
    out: dict[int, dict] = {}
    for r in rows:
        aid = int(r["aid"])
        ftp = _parse_bq_date(r.get("ftp_date"))
        sen_days = (report_date - ftp).days if ftp else None
        out[aid] = {
            "net_lt": float(r.get("net_lt") or 0),
            "net_30d": float(r.get("net_30d") or 0),
            "net_60d": float(r.get("net_60d") or 0),
            "purchased_30d": float(r.get("purchased_30d") or 0),
            "hold": float(r["hold_lt"]) if r.get("hold_lt") is not None else None,
            "reg_date": _parse_bq_date(r.get("reg_date")),
            "ftp_date": ftp,
            "seniority_days": sen_days,
            "invite_status": invite_status_label(r.get("locked"), r.get("lock_reason")),
        }
    return out


def apply_enrich_to_players(players: list[Player], metrics: dict[int, dict], report_date: date) -> None:
    for p in players:
        m = metrics.get(p.aid, {})
        p.net_lt = round(m.get("net_lt", p.net_lt), 2)
        p.net_30d = round(m.get("net_30d", 0), 2)
        p.net_60d = round(m.get("net_60d", 0), 2)
        p.purchased_30d = round(m.get("purchased_30d", 0), 2)
        if m.get("hold") is not None:
            p.hold = float(m["hold"])
        p.reg_date = m.get("reg_date") or p.reg_date
        p.ftp_date = m.get("ftp_date") or p.ftp_date
        p.seniority_days = m.get("seniority_days")
        if p.seniority_days is None and p.ftp_date:
            p.seniority_days = (report_date - p.ftp_date).days
        p.seniority_group = seniority_group(p.seniority_days)
        p.invite_status = m.get("invite_status", "Active")


def _norm_header(val) -> str:
    return str(val or "").lower().replace("\n", " ").strip()


def _find_col(headers: dict[int, str], *needles: str) -> int | None:
    for col, h in headers.items():
        if all(n in h for n in needles):
            return col
    return None


def load_players(source: Path) -> list[Player]:
    wb = load_workbook(source, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if "top 50" in name.lower() or "vip event" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    hdr_row = None
    for r in range(1, 15):
        if _norm_header(ws.cell(r, 1).value) == "rank":
            hdr_row = r
            break
    if hdr_row is None:
        hdr_row = 8

    raw = {c: _norm_header(ws.cell(hdr_row, c).value) for c in range(1, 20)}
    headers = {c: h for c, h in raw.items() if h}

    col_rank = _find_col(headers, "rank") or 1
    col_aid = _find_col(headers, "aid") or _find_col(headers, "account") or 2
    col_name = _find_col(headers, "full", "name") or _find_col(headers, "name") or 3
    col_age = _find_col(headers, "age") or 4
    col_state = _find_col(headers, "state") or 5
    col_agent = _find_col(headers, "agent") or 6
    col_lt = _find_col(headers, "lifetime") or _find_col(headers, "net", "purchase") or 7
    col_hold = _find_col(headers, "hold") or 10

    players: list[Player] = []
    row = hdr_row + 1
    while row <= hdr_row + 55:
        rank = ws.cell(row, col_rank).value
        aid = ws.cell(row, col_aid).value
        if rank is None or aid is None:
            if str(ws.cell(row, col_rank).value or "").upper() in ("AVG", "TOTAL"):
                break
            row += 1
            continue
        try:
            rank_i = int(rank)
            aid_i = int(aid)
        except (TypeError, ValueError):
            row += 1
            continue
        if rank_i < 1 or rank_i > 50:
            row += 1
            continue

        hold_val = ws.cell(row, col_hold).value
        players.append(
            Player(
                rank=rank_i,
                aid=aid_i,
                name=str(ws.cell(row, col_name).value or ""),
                age=ws.cell(row, col_age).value,
                state=str(ws.cell(row, col_state).value or ""),
                agent=str(ws.cell(row, col_agent).value or ""),
                net_lt=float(ws.cell(row, col_lt).value or 0),
                hold=float(hold_val) if hold_val is not None else None,
            )
        )
        row += 1

    players.sort(key=lambda p: p.rank)
    return players[:50]


def fetch_locked_aids(aids: list[int]) -> set[int]:
    if not aids:
        return set()
    in_list = ", ".join(str(a) for a in aids)
    sql = f"""
SELECT id AS aid
FROM `{PROJECT_ID}.transactional_data.uam_accounts`
WHERE id IN ({in_list})
  AND (locked = TRUE OR LOWER(COALESCE(lock_reason, '')) LIKE '%exclusion%')
"""
    return {int(r["aid"]) for r in run_query(get_client(), sql)}


def state_breakdown(players: list[Player]) -> list[tuple[str, int, float]]:
    counts = Counter(p.state for p in players if p.state)
    total = sum(counts.values()) or 1
    return [(name, n, n / total) for name, n in counts.most_common()]


def seniority_breakdown(players: list[Player]) -> list[tuple[str, int, float]]:
    counts = Counter(p.seniority_group for p in players if p.seniority_group and p.seniority_group != "—")
    total = sum(counts.values()) or 1
    ordered = [(g, counts[g], counts[g] / total) for g in SENIORITY_BUCKET_ORDER if counts.get(g)]
    return ordered


def cohort_summary(players: list[Player], report_date: date) -> dict:
    states = state_breakdown(players)
    top_name, top_n, top_pct = states[0] if states else ("—", 0, 0.0)
    holds = [p.hold for p in players if p.hold is not None]
    ages = [float(p.age) for p in players if p.age is not None]
    med_sen_days = median_seniority_days(players)
    return {
        "as_of": report_date,
        "count": len(players),
        "sum_np_lt": sum(p.net_lt for p in players),
        "sum_np_30d": sum(p.net_30d for p in players),
        "sum_np_60d": sum(p.net_60d for p in players),
        "sum_purchased_30d": sum(p.purchased_30d for p in players),
        "avg_np_lt": sum(p.net_lt for p in players) / len(players),
        "avg_np_30d": sum(p.net_30d for p in players) / len(players),
        "avg_np_60d": sum(p.net_60d for p in players) / len(players),
        "avg_hold": sum(holds) / len(holds) if holds else 0,
        "median_age": sorted(ages)[len(ages) // 2] if ages else 0,
        "median_seniority_days": med_sen_days,
        "median_seniority_group": seniority_group(med_sen_days),
        "top_state": top_name,
        "top_state_n": top_n,
        "top_state_pct": top_pct,
        "states": states,
        "seniority": seniority_breakdown(players),
    }


def median_seniority_kpi(players: list[Player]) -> str:
    days = median_seniority_days(players)
    if days is None:
        return "—"
    return f"{days:,} d\n{seniority_group(days)}"


def top_state_label(players: list[Player]) -> str:
    counts = Counter(p.state for p in players if p.state)
    if not counts:
        return "—"
    name, n = counts.most_common(1)[0]
    return f"{name}  ·  {n} of 50  ({n / 50:.0%})"


def add_seniority_share_section(ws, players: list[Player], start_row: int) -> int:
    """FTP tenure mix by database seniority bucket."""
    white = PatternFill("solid", fgColor=WHITE)
    buckets = seniority_breakdown(players)
    section_fill = PatternFill("solid", fgColor=NAVY)
    hdr_fill = PatternFill("solid", fgColor=HEADER_BG)
    alt_fill = PatternFill("solid", fgColor=SURFACE_ALT)

    merge_style(
        ws, start_row, 1, start_row, 4,
        "  TENURE — DAYS FROM FTP (COHORT MIX)",
        Font(name="Calibri", size=11, bold=True, color=WHITE),
        section_fill,
        Alignment(horizontal="left", vertical="center", indent=1),
    )
    ws.row_dimensions[start_row].height = 22

    hr = start_row + 1
    for col, label in ((1, "Seniority group"), (2, "# Players"), (3, "% of 50")):
        set_cell(
            ws, hr, col, label, FONT_TH, hdr_fill,
            Alignment(horizontal="center" if col > 1 else "left", vertical="center"),
            border=BORDER,
        )
    ws.row_dimensions[hr].height = 24

    for i, (name, n, pct) in enumerate(buckets):
        row = hr + 1 + i
        row_fill = alt_fill if i % 2 else white
        set_cell(ws, row, 1, name, FONT_BODY, row_fill, Alignment(horizontal="left"), border=BORDER)
        set_cell(ws, row, 2, n, FONT_BODY, row_fill, Alignment(horizontal="center"), "0", BORDER)
        set_cell(ws, row, 3, pct, FONT_BODY, row_fill, Alignment(horizontal="center"), PCT_FMT, BORDER)

    total_row = hr + 1 + len(buckets)
    total_fill = PatternFill("solid", fgColor=TOTAL_BG)
    set_cell(ws, total_row, 1, "TOTAL", FONT_TOTAL, total_fill, Alignment(horizontal="left"), border=BORDER)
    set_cell(ws, total_row, 2, 50, FONT_TOTAL, total_fill, Alignment(horizontal="center"), "0", BORDER)
    set_cell(ws, total_row, 3, 1.0, FONT_TOTAL, total_fill, Alignment(horizontal="center"), PCT_FMT, BORDER)
    ws.row_dimensions[total_row].height = 22
    return total_row + 2


def add_state_share_section(ws, players: list[Player], start_row: int) -> int:
    """Write geography table; return next free row."""
    white = PatternFill("solid", fgColor=WHITE)
    states = state_breakdown(players)
    section_fill = PatternFill("solid", fgColor=NAVY)
    hdr_fill = PatternFill("solid", fgColor=HEADER_BG)
    alt_fill = PatternFill("solid", fgColor=SURFACE_ALT)

    merge_style(
        ws, start_row, 1, start_row, 4,
        "  GEOGRAPHY — STATE SHARE",
        Font(name="Calibri", size=11, bold=True, color=WHITE),
        section_fill,
        Alignment(horizontal="left", vertical="center", indent=1),
    )
    ws.row_dimensions[start_row].height = 22

    hr = start_row + 1
    for col, label in ((1, "State"), (2, "# Players"), (3, "% of 50")):
        set_cell(
            ws, hr, col, label, FONT_TH, hdr_fill,
            Alignment(horizontal="center" if col > 1 else "left", vertical="center"),
            border=BORDER,
        )
    ws.row_dimensions[hr].height = 24

    for i, (name, n, pct) in enumerate(states):
        row = hr + 1 + i
        row_fill = alt_fill if i % 2 else white
        set_cell(ws, row, 1, name, FONT_BODY, row_fill, Alignment(horizontal="left"), border=BORDER)
        set_cell(ws, row, 2, n, FONT_BODY, row_fill, Alignment(horizontal="center"), "0", BORDER)
        set_cell(ws, row, 3, pct, FONT_BODY, row_fill, Alignment(horizontal="center"), PCT_FMT, BORDER)

    total_row = hr + 1 + len(states)
    total_fill = PatternFill("solid", fgColor=TOTAL_BG)
    set_cell(ws, total_row, 1, "TOTAL", FONT_TOTAL, total_fill, Alignment(horizontal="left"), border=BORDER)
    set_cell(ws, total_row, 2, 50, FONT_TOTAL, total_fill, Alignment(horizontal="center"), "0", BORDER)
    set_cell(ws, total_row, 3, 1.0, FONT_TOTAL, total_fill, Alignment(horizontal="center"), PCT_FMT, BORDER)
    ws.row_dimensions[total_row].height = 22

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 15, 18)
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    return total_row + 2


def set_cell(ws, row, col, value=None, font=None, fill=None, align=None, fmt=None, border=None):
    c = ws.cell(row, col)
    if value is not None:
        c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if fmt:
        c.number_format = fmt
    if border:
        c.border = border
    return c


def merge_style(ws, r1, c1, r2, c2, value=None, font=None, fill=None, align=None, fmt=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    set_cell(ws, r1, c1, value, font, fill, align, fmt)


def build_workbook(players: list[Player], report_date: date, active_count: int, summary: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Detail — Full Data"
    ws.sheet_view.showGridLines = False

    navy_fill = PatternFill("solid", fgColor=NAVY)
    orange_light = PatternFill("solid", fgColor=ORANGE_LIGHT)
    white = PatternFill("solid", fgColor=WHITE)
    warn_fill = PatternFill("solid", fgColor=WARN_FILL_COLOR)
    metric_font = FONT_KPI_VAL
    med_sen_kpi = median_seniority_kpi(players)
    med_sen_days = median_seniority_days(players)

    # ── Header band ──────────────────────────────────────────────
    merge_style(
        ws, 1, 1, 2, LAST_COL,
        "VIP Event",
        FONT_H1,
        navy_fill,
        Alignment(horizontal="left", vertical="center", indent=1),
    )
    merge_style(
        ws, 3, 1, 3, LAST_COL,
        f"Top 50 Players  ·  Executive Brief",
        FONT_H2,
        navy_fill,
        Alignment(horizontal="left", vertical="center", indent=1),
    )
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 6
    ws.row_dimensions[3].height = 22

    prepared = date.today().strftime("%b %d, %Y")
    as_of = report_date.strftime("%b %d, %Y")
    merge_style(
        ws, 4, 1, 4, LAST_COL,
        f"Data as of {as_of}   |   {active_count} active invitees (locked / self-exclusion excluded)   |   Managed Elite book",
        FONT_META,
        PatternFill("solid", fgColor=NAVY_MID),
        Alignment(horizontal="left", vertical="center", indent=1),
    )
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 10

    # ── KPI cards (row 6–7) ─────────────────────────────────────
    # 15 cols: four 2-col NP cards + hold + top state + median age + median seniority
    kpi_pairs = [
        (1, 2, "=AVERAGE(J10:J59)", MONEY_FMT, "AVG NET PURCHASE\nLIFETIME", metric_font),
        (3, 4, "=AVERAGE(K10:K59)", MONEY_FMT, "AVG NET PURCHASE\n30 DAYS", metric_font),
        (5, 6, "=AVERAGE(L10:L59)", MONEY_FMT, "AVG NET PURCHASE\n60 DAYS", metric_font),
        (7, 7, "=AVERAGE(N10:N59)", PCT_FMT, "AVG HOLD %\nLIFETIME", metric_font),
        (8, 9, top_state_label(players), None, "TOP STATE\nBY COUNT", Font(name="Calibri", size=12, bold=True, color=ORANGE[2:])),
        (10, 11, "=MEDIAN(D10:D59)", "0", "MEDIAN\nAGE", metric_font),
        (12, 13, med_sen_kpi, None, "MEDIAN\nDAYS FROM FTP", Font(name="Calibri", size=11, bold=True, color=ORANGE[2:])),
    ]
    card_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lbl_align = Alignment(horizontal="center", vertical="top", wrap_text=True)

    for c1, c2, formula, fmt, label, val_font in kpi_pairs:
        if c1 == c2:
            set_cell(ws, 6, c1, formula, val_font, orange_light, card_align, fmt, BORDER)
            set_cell(ws, 7, c1, label, FONT_KPI_LBL, white, lbl_align, border=BORDER)
        else:
            merge_style(ws, 6, c1, 6, c2, formula, val_font, orange_light, card_align, fmt)
            merge_style(ws, 7, c1, 7, c2, label, FONT_KPI_LBL, white, lbl_align)
            for c in range(c1, c2 + 1):
                set_cell(ws, 6, c, border=BORDER)
                set_cell(ws, 7, c, border=BORDER)

    ws.row_dimensions[6].height = 38
    ws.row_dimensions[7].height = 30
    ws.row_dimensions[8].height = 14

    # Orange accent line above KPI cards
    for c in range(1, LAST_COL + 1):
        set_cell(ws, 5, c, fill=PatternFill("solid", fgColor=ORANGE))

    set_cell(
        ws, 8, 1, "Player roster", FONT_SECTION, white,
        Alignment(horizontal="left", vertical="bottom"),
        border=BORDER_BOTTOM_ACCENT,
    )
    for c in range(2, LAST_COL + 1):
        set_cell(ws, 8, c, border=BORDER_BOTTOM_ACCENT)

    # ── Table headers ───────────────────────────────────────────
    headers = [
        "Rank", "AID", "Player", "Age", "State", "Agent",
        "Reg Date", "Days from FTP", "Seniority group",
        "Net Purchase\nLifetime", "Net Purchase\n30d", "Net Purchase\n60d",
        "Purchased\n30d", "Hold %", "Status",
    ]
    hdr_fill = PatternFill("solid", fgColor=HEADER_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, text in enumerate(headers, 1):
        set_cell(ws, HDR_ROW, col, text, FONT_TH, hdr_fill, hdr_align, border=BORDER)
    ws.row_dimensions[HDR_ROW].height = 36

    # ── Data rows ─────────────────────────────────────────────────
    for i, p in enumerate(players):
        row = DATA_START + i
        alt = i % 2 == 1
        base_fill = PatternFill("solid", fgColor=SURFACE_ALT if alt else WHITE)
        row_fill = warn_fill if p.invite_status != "Active" else base_fill

        values = [
            (COL_RANK, p.rank, FONT_RANK, "center", None),
            (COL_AID, p.aid, FONT_BODY, "left", None),
            (COL_NAME, p.name, FONT_NAME, "left", None),
            (COL_AGE, p.age, FONT_BODY, "center", "0"),
            (COL_STATE, p.state, FONT_BODY, "left", None),
            (COL_AGENT, p.agent, FONT_BODY, "left", None),
            (COL_REG, p.reg_date, FONT_BODY, "center", DATE_FMT),
            (COL_SEN_DAYS, p.seniority_days, FONT_BODY, "center", "#,##0"),
            (COL_SEN_GROUP, p.seniority_group, FONT_BODY, "center", None),
            (COL_NP_LT, p.net_lt, FONT_MONEY, "right", MONEY_FMT),
            (COL_NP_30, p.net_30d, FONT_MONEY, "right", MONEY_FMT),
            (COL_NP_60, p.net_60d, FONT_MONEY, "right", MONEY_FMT),
            (COL_PURCH_30, p.purchased_30d, FONT_MONEY, "right", MONEY_FMT),
            (COL_HOLD, p.hold, FONT_MONEY, "right", PCT_FMT),
            (COL_STATUS, p.invite_status, FONT_BODY, "center", None),
        ]
        for col, val, font, halign, fmt in values:
            set_cell(
                ws, row, col, val, font, row_fill,
                Alignment(horizontal=halign, vertical="center"), fmt, BORDER,
            )

    # ── Totals row ────────────────────────────────────────────────
    total_fill = PatternFill("solid", fgColor=TOTAL_BG)
    set_cell(ws, TOTAL_ROW, COL_RANK, "TOTAL", FONT_TOTAL, total_fill, Alignment(horizontal="left"), border=BORDER)
    set_cell(ws, TOTAL_ROW, COL_NAME, "50 players", FONT_TOTAL, total_fill, Alignment(horizontal="left"), border=BORDER)
    set_cell(ws, TOTAL_ROW, COL_SEN_DAYS, med_sen_days, FONT_TOTAL, total_fill,
             Alignment(horizontal="center"), "#,##0", BORDER)
    set_cell(ws, TOTAL_ROW, COL_SEN_GROUP, seniority_group(med_sen_days), FONT_TOTAL, total_fill,
             Alignment(horizontal="center"), border=BORDER)
    for col in (COL_AID, COL_AGE, COL_STATE, COL_AGENT, COL_REG, COL_STATUS):
        set_cell(ws, TOTAL_ROW, col, border=BORDER, fill=total_fill)

    for col, formula, fmt in (
        (COL_AGE, "=SUM(D10:D59)", "0"),
        (COL_NP_LT, "=SUM(J10:J59)", MONEY_FMT),
        (COL_NP_30, "=SUM(K10:K59)", MONEY_FMT),
        (COL_NP_60, "=SUM(L10:L59)", MONEY_FMT),
        (COL_PURCH_30, "=SUM(M10:M59)", MONEY_FMT),
        (COL_HOLD, "=AVERAGE(N10:N59)", PCT_FMT),
    ):
        set_cell(
            ws, TOTAL_ROW, col, formula, FONT_TOTAL, total_fill,
            Alignment(horizontal="right", vertical="center"), fmt, BORDER,
        )

    ws.row_dimensions[TOTAL_ROW].height = 24

    merge_style(
        ws, TOTAL_ROW + 2, 1, TOTAL_ROW + 2, LAST_COL,
        "Net purchase = purchased − redeemed − chargebacks − refunds  ·  Days from FTP = DATE_DIFF(as-of, ftp_date)  ·  Seniority group = database cohort  ·  Managed Elite book",
        Font(name="Calibri", size=9, italic=True, color=MUTED[2:]),
        white,
        Alignment(horizontal="left", indent=1),
    )
    ws.row_dimensions[TOTAL_ROW + 2].height = 16

    next_row = add_seniority_share_section(ws, players, TOTAL_ROW + 4)
    add_state_share_section(ws, players, next_row)

    # ── Column widths & freeze ────────────────────────────────────
    widths = {
        1: 7, 2: 13, 3: 22, 4: 7, 5: 14, 6: 12,
        7: 12, 8: 11, 9: 14, 10: 14, 11: 12, 12: 12, 13: 12, 14: 10, 15: 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = f"A{DATA_START}"
    ws.print_title_rows = f"{HDR_ROW}:{HDR_ROW}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    build_presentation_slides(wb, players, summary, report_date, active_count)
    return wb


def generate(source: Path, report_date: date, output: Path | None = None) -> tuple[Path, dict]:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    players = load_players(source)
    if len(players) < 50:
        raise ValueError(f"Expected 50 players, found {len(players)} in {source}")

    metrics = fetch_enrich_metrics([p.aid for p in players], report_date)
    apply_enrich_to_players(players, metrics, report_date)

    locked = fetch_locked_aids([p.aid for p in players])
    active_count = len(players) - len(locked)
    if locked:
        print(f"Warning: {len(locked)} invitee(s) locked or self-excluded: {sorted(locked)}")

    summary = cohort_summary(players, report_date)
    wb = build_workbook(players, report_date, active_count, summary)
    out = output or EXPORT_DIR / "VIP Event - Top 50 - Management Brief.xlsx"

    try:
        wb.save(out)
    except PermissionError:
        out = EXPORT_DIR / f"VIP-Event-Top50-Management-Brief-{report_date.isoformat()}-new.xlsx"
        wb.save(out)

    print(f"Wrote: {out}")
    print(f"Players: {len(players)}  |  Active: {active_count}  |  NP 30d total: ${summary['sum_np_30d']:,.0f}")
    return out, summary


def main() -> None:
    parser = argparse.ArgumentParser(description="VIP Event Top 50 management brief")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument(
        "--date",
        type=lambda s: date.fromisoformat(s),
        default=date.today() - timedelta(days=1),
    )
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()
    path, _summary = generate(args.source, args.date, args.output)
    try:
        import os
        os.startfile(path)  # noqa: S606 — open in Excel on Windows
    except OSError:
        pass


if __name__ == "__main__":
    main()
