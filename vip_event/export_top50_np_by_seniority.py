"""
Export Vegas Top 50 net purchase summary by seniority — group totals + avg per player.

Usage:
  python vip_event/export_top50_np_by_seniority.py
  python vip_event/export_top50_np_by_seniority.py --date YYYY-MM-DD
"""
from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "vip_event"))

from generate_top50_management_brief import (  # noqa: E402
    DEFAULT_SOURCE,
    fetch_enrich_metrics,
    load_players,
    seniority_group,
)

MODULE_DIR = Path(__file__).resolve().parent
EXPORT_DIR = MODULE_DIR / "exports"

GROUP_ORDER = ("First 31-90d", "First 91-180", "First 181-365", "+1Y")

NAVY = "FF1B2A4A"
WHITE = "FFFFFFFF"
INK = "FF1A1A2E"
RULE = "FFE2E8F0"
HEADER_BG = "FFE8EEF5"
TOTAL_BG = "FFD4E0ED"
ALT_ROW = "FFF7F9FC"
MONEY_FMT = '"$"#,##0'

SIDE = Side(style="thin", color=RULE)
BORDER = Border(left=SIDE, right=SIDE, top=SIDE, bottom=SIDE)


@dataclass
class GroupRow:
    seniority: str
    players: int
    np_30d_total: float
    np_30d_avg: float
    np_60d_total: float
    np_60d_avg: float
    np_lt_total: float
    np_lt_avg: float


def _avg(vals: list[float]) -> float:
    return sum(vals) / len(vals) if vals else 0.0


def build_group_rows(players: list, metrics: dict[int, dict], report_date: date) -> list[GroupRow]:
    buckets: dict[str, list[dict]] = defaultdict(list)
    for p in players:
        m = metrics.get(p.aid, {})
        sen = m.get("seniority_days")
        if sen is None and m.get("ftp_date"):
            sen = (report_date - m["ftp_date"]).days
        grp = seniority_group(sen)
        buckets[grp].append(
            {
                "np30": float(m.get("net_30d") or 0),
                "np60": float(m.get("net_60d") or 0),
                "nplt": float(m.get("net_lt") or 0),
            }
        )

    rows: list[GroupRow] = []
    for grp in GROUP_ORDER:
        if grp not in buckets:
            continue
        items = buckets[grp]
        n = len(items)
        rows.append(
            GroupRow(
                seniority=grp,
                players=n,
                np_30d_total=sum(x["np30"] for x in items),
                np_30d_avg=_avg([x["np30"] for x in items]),
                np_60d_total=sum(x["np60"] for x in items),
                np_60d_avg=_avg([x["np60"] for x in items]),
                np_lt_total=sum(x["nplt"] for x in items),
                np_lt_avg=_avg([x["nplt"] for x in items]),
            )
        )

    all_items = [x for items in buckets.values() for x in items]
    rows.append(
        GroupRow(
            seniority="ALL PLAYERS",
            players=len(all_items),
            np_30d_total=sum(x["np30"] for x in all_items),
            np_30d_avg=_avg([x["np30"] for x in all_items]),
            np_60d_total=sum(x["np60"] for x in all_items),
            np_60d_avg=_avg([x["np60"] for x in all_items]),
            np_lt_total=sum(x["nplt"] for x in all_items),
            np_lt_avg=_avg([x["nplt"] for x in all_items]),
        )
    )
    return rows


def write_csv(path: Path, rows: list[GroupRow], report_date: date) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([f"Vegas Top 50 — Net purchase by seniority (as of {report_date.isoformat()})"])
        w.writerow([])
        w.writerow(
            [
                "Seniority",
                "Players",
                "NP 30d Total",
                "NP 30d Avg PP",
                "NP 60d Total",
                "NP 60d Avg PP",
                "NP LT Total",
                "NP LT Avg PP",
            ]
        )
        for r in rows:
            w.writerow(
                [
                    r.seniority,
                    r.players,
                    round(r.np_30d_total, 2),
                    round(r.np_30d_avg, 2),
                    round(r.np_60d_total, 2),
                    round(r.np_60d_avg, 2),
                    round(r.np_lt_total, 2),
                    round(r.np_lt_avg, 2),
                ]
            )


def _set_cell(ws, row: int, col: int, value, *, font, fill=None, align="left", fmt=None, bold=False):
    c = ws.cell(row, col, value)
    c.font = font
    if fill:
        c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    return c


def write_xlsx(path: Path, rows: list[GroupRow], report_date: date) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "NP by Seniority"
    ws.sheet_view.showGridLines = False

    font_title = Font(name="Calibri", size=16, bold=True, color=NAVY[2:])
    font_sub = Font(name="Calibri", size=10, color="FF6B7C93")
    font_th = Font(name="Calibri", size=10, bold=True, color=NAVY[2:])
    font_body = Font(name="Calibri", size=11, color=INK[2:])
    font_total = Font(name="Calibri", size=11, bold=True, color=NAVY[2:])
    fill_hdr = PatternFill("solid", fgColor=HEADER_BG)
    fill_total = PatternFill("solid", fgColor=TOTAL_BG)
    fill_alt = PatternFill("solid", fgColor=ALT_ROW)

    ws.merge_cells("A1:H1")
    ws["A1"] = "Vegas Top 50 — Net Purchase by Seniority"
    ws["A1"].font = font_title
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Data as of {report_date.strftime('%d %b %Y')} · Group total + average per player (PP)"
    ws["A2"].font = font_sub

    hdr_row = 4
    headers = [
        ("Seniority", "left", None),
        ("Players", "center", None),
        ("NP 30d Total", "right", MONEY_FMT),
        ("NP 30d Avg PP", "right", MONEY_FMT),
        ("NP 60d Total", "right", MONEY_FMT),
        ("NP 60d Avg PP", "right", MONEY_FMT),
        ("NP LT Total", "right", MONEY_FMT),
        ("NP LT Avg PP", "right", MONEY_FMT),
    ]
    for col, (label, align, fmt) in enumerate(headers, start=1):
        _set_cell(ws, hdr_row, col, label, font=font_th, fill=fill_hdr, align=align, fmt=fmt)

    for i, r in enumerate(rows):
        row_num = hdr_row + 1 + i
        is_total = r.seniority == "ALL PLAYERS"
        fill = fill_total if is_total else (fill_alt if i % 2 else None)
        font = font_total if is_total else font_body
        vals = [
            (r.seniority, "left", None),
            (r.players, "center", None),
            (r.np_30d_total, "right", MONEY_FMT),
            (r.np_30d_avg, "right", MONEY_FMT),
            (r.np_60d_total, "right", MONEY_FMT),
            (r.np_60d_avg, "right", MONEY_FMT),
            (r.np_lt_total, "right", MONEY_FMT),
            (r.np_lt_avg, "right", MONEY_FMT),
        ]
        for col, (val, align, fmt) in enumerate(vals, start=1):
            _set_cell(ws, row_num, col, val, font=font, fill=fill, align=align, fmt=fmt)

    widths = (18, 9, 14, 14, 14, 14, 14, 14)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    wb.save(path)


def _fmt_money(n: float) -> str:
    return f"${n:,.0f}"


def write_html(path: Path, rows: list[GroupRow], report_date: date) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    body_rows = []
    for r in rows:
        cls = ' class="total"' if r.seniority == "ALL PLAYERS" else ""
        body_rows.append(
            f"<tr{cls}>"
            f"<td>{r.seniority}</td>"
            f"<td class='num'>{r.players}</td>"
            f"<td class='money'>{_fmt_money(r.np_30d_total)}</td>"
            f"<td class='money avg'>{_fmt_money(r.np_30d_avg)}</td>"
            f"<td class='money'>{_fmt_money(r.np_60d_total)}</td>"
            f"<td class='money avg'>{_fmt_money(r.np_60d_avg)}</td>"
            f"<td class='money'>{_fmt_money(r.np_lt_total)}</td>"
            f"<td class='money avg'>{_fmt_money(r.np_lt_avg)}</td>"
            "</tr>"
        )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <title>Vegas Top 50 — NP by Seniority</title>
  <style>
    body {{ font-family: "Segoe UI", system-ui, sans-serif; margin: 32px; color: #1a1a2e; background: #f7f9fc; }}
    h1 {{ margin: 0 0 4px; font-size: 22px; color: #1b2a4a; }}
    .sub {{ color: #6b7c93; font-size: 13px; margin-bottom: 20px; }}
    table {{ border-collapse: collapse; width: 100%; max-width: 1100px; background: #fff; box-shadow: 0 1px 4px rgba(0,0,0,.08); }}
    th, td {{ border: 1px solid #e2e8f0; padding: 10px 12px; font-size: 13px; }}
    th {{ background: #e8eef5; color: #1b2a4a; font-weight: 600; text-align: left; }}
    th.group {{ text-align: center; border-bottom: 2px solid #3d6b9e; }}
    th.subcol {{ text-align: right; font-weight: 500; font-size: 12px; }}
    td.num {{ text-align: center; }}
    td.money {{ text-align: right; font-variant-numeric: tabular-nums; }}
    td.avg {{ color: #3d6b9e; font-weight: 600; }}
    tr:nth-child(even):not(.total) td {{ background: #f7f9fc; }}
    tr.total td {{ background: #d4e0ed; font-weight: 700; }}
    .legend {{ margin-top: 12px; font-size: 12px; color: #6b7c93; }}
  </style>
</head>
<body>
  <h1>Vegas Top 50 — Net Purchase by Seniority</h1>
  <p class="sub">Data as of {report_date.strftime("%d %b %Y")} · Blue columns = average per player (PP)</p>
  <table>
    <thead>
      <tr>
        <th rowspan="2">Seniority</th>
        <th rowspan="2" style="text-align:center">Players</th>
        <th colspan="2" class="group">NP 30d</th>
        <th colspan="2" class="group">NP 60d</th>
        <th colspan="2" class="group">NP Lifetime</th>
      </tr>
      <tr>
        <th class="subcol">Group total</th>
        <th class="subcol">Avg PP</th>
        <th class="subcol">Group total</th>
        <th class="subcol">Avg PP</th>
        <th class="subcol">Group total</th>
        <th class="subcol">Avg PP</th>
      </tr>
    </thead>
    <tbody>
      {"".join(body_rows)}
    </tbody>
  </table>
  <p class="legend">NP = purchased − redeemed − chargebacks − refunds (account-day grain, Elite book).</p>
</body>
</html>"""
    path.write_text(html, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", type=lambda s: date.fromisoformat(s))
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    args = parser.parse_args()
    report_date = args.date or (date.today() - timedelta(days=1))

    players = load_players(args.source)
    metrics = fetch_enrich_metrics([p.aid for p in players], report_date)
    rows = build_group_rows(players, metrics, report_date)

    stem = f"vegas-top50-np-by-seniority-{report_date.isoformat()}"
    csv_path = EXPORT_DIR / f"{stem}.csv"
    xlsx_path = EXPORT_DIR / f"{stem}.xlsx"
    html_path = EXPORT_DIR / f"{stem}.html"

    write_csv(csv_path, rows, report_date)
    write_xlsx(xlsx_path, rows, report_date)
    write_html(html_path, rows, report_date)

    print(f"Report date: {report_date.isoformat()}")
    print(f"CSV:   {csv_path}")
    print(f"Excel: {xlsx_path}")
    print(f"HTML:  {html_path}")


if __name__ == "__main__":
    main()
