from openpyxl import load_workbook
from pathlib import Path

p = Path(r"c:\Users\Owner\Downloads\Elite\vip_event\data\vip-event-top50-players.xlsx")
wb = load_workbook(p, data_only=False)
ws = wb["VIP Event - Top 50 Players "]
formulas = []
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            formulas.append(f"{cell.coordinate}: {cell.value}")
Path(r"c:\Users\Owner\Downloads\Elite\vip_event\data\xlsx-formulas.txt").write_text(
    "\n".join(formulas), encoding="utf-8"
)
print(len(formulas), "formulas")
