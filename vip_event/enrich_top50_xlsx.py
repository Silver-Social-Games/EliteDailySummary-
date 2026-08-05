"""
Enrich VIP Event Top 50 Players xlsx — simple player table + summary KPIs.

Usage:
  python vip_event/enrich_top50_xlsx.py
  python vip_event/enrich_top50_xlsx.py --source "path/to/file.xlsx"
"""
from __future__ import annotations

import argparse
import shutil
import sys
from collections import Counter
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from daily_summary.generate_daily_elite_summary import PROJECT_ID, get_client, run_query

MODULE_DIR = Path(__file__).resolve().parent
DATA_DIR = MODULE_DIR / "data"

DEFAULT_SOURCE = Path(
    r"c:\Users\Owner\OneDrive - Silver Social Games\Desktop\VIP\VIP Event"
    r"\VIP Event - Top 50 Players .xlsx"
)
SHEET_NAME = "VIP Event - Top 50 Players "

HDR_ROW = 8
DATA_START = 9
DATA_END = 58
TOTAL_ROW = 59

COL_RANK, COL_AID, COL_NAME = 1, 2, 3
COL_AGE, COL_STATE, COL_AGENT = 4, 5, 6
COL_LT, COL_NP30, COL_NP60, COL_HOLD = 7, 8, 9, 10
LAST_COL = COL_HOLD

# Theme
NAVY = "FF2C3E6E"
STEEL = "FF4A6FA5"
METRIC_BG = "FFEBF0F8"
HEADER_FILL = "FFD4E4F5"
ALT_FILL = "FFF0F5FB"
TOTAL_FILL = "FFC5D9EE"
WHITE = "FFFFFFFF"
THIN = Side(style="thin", color="FFD0D8E4")

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color="FFFFFFFF")
FONT_SECTION = Font(name="Calibri", size=10, bold=True, color=STEEL[2:])
FONT_KPI_VALUE = Font(name="Calibri", size=20, bold=True, color="FF1A1A2E")
FONT_KPI_LABEL = Font(name="Calibri", size=9, bold=True, color="FF5A6A7E")
FONT_TABLE_HDR = Font(name="Calibri", size=10, bold=True, color="FF1A1A2E")
FONT_NAME = Font(name="Calibri", size=11, bold=True)
FONT_BODY = Font(name="Calibri", size=11)
FONT_RANK = Font(name="Calibri", size=11, bold=True)
FONT_MONEY = Font(name="Calibri", size=11, bold=True)
FONT_TOTAL = Font(name="Calibri", size=11, bold=True)

MONEY_FMT = '"$"#,##0'
PCT_FMT = "0.0%"

KPI_COLS = (1, 4, 7, 10, 13, 16)  # A, D, G, J, M, P
KPI_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
KPI_LABEL_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TABLE_HEADERS = {
    COL_RANK: "Rank",
    COL_AID: "AID",
    COL_NAME: "Full Name",
    COL_AGE: "Age",
    COL_STATE: "State",
    COL_AGENT: "Agent",
    COL_LT: "Net Purchase\nLifetime",
    COL_NP30: "Net Purchase\n30d",
    COL_NP60: "Net Purchase\n60d",
    COL_HOLD: "Hold %",
}


def copy_workbook_source(src: Path) -> Path:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    dst = DATA_DIR / "vip-event-top50-players.xlsx"
    shutil.copy2(src, dst)
    return dst


def build_np_sql(aids: list[int], report_date: date) -> str:
    rd = report_date.isoformat()
    d30 = (report_date - timedelta(days=29)).isoformat()
    d60 = (report_date - timedelta(days=59)).isoformat()
    in_list = ", ".join(str(a) for a in aids)
    return f"""
WITH daily AS (
  SELECT
    account_id AS aid,
    date,
    SUM(
      CAST(purchased AS FLOAT64)
      - CAST(redeemed AS FLOAT64)
      - CAST(chargeback AS FLOAT64)
      - CAST(refunds AS FLOAT64)
    ) AS net_purchases
  FROM `{PROJECT_ID}.jackpota_agg.daily_player_revenue_kpis`
  WHERE account_id IN ({in_list})
  GROUP BY account_id, date
)
SELECT
  aid,
  SUM(IF(date BETWEEN DATE '{d30}' AND DATE '{rd}', net_purchases, 0)) AS net_purchase_30d,
  SUM(IF(date BETWEEN DATE '{d60}' AND DATE '{rd}', net_purchases, 0)) AS net_purchase_60d
FROM daily
GROUP BY aid
"""


def fetch_net_purchases(aids: list[int], report_date: date) -> dict[int, dict[str, float]]:
    if not aids:
        return {}
    rows = run_query(get_client(), build_np_sql(aids, report_date))
    return {
        int(r["aid"]): {
            "net_purchase_30d": float(r.get("net_purchase_30d") or 0),
            "net_purchase_60d": float(r.get("net_purchase_60d") or 0),
        }
        for r in rows
    }


def fill(cell, value=None, font=None, fill_=None, align=None, number_format=None, border=None) -> None:
    if value is not None:
        cell.value = value
    if font is not None:
        cell.font = font
    if fill_ is not None:
        cell.fill = fill_
    if align is not None:
        cell.alignment = align
    if number_format is not None:
        cell.number_format = number_format
    if border is not None:
        cell.border = border


def has_np_columns(ws) -> bool:
    h8 = ws.cell(HDR_ROW, COL_NP30).value
    return bool(h8 and "Net Purchase" in str(h8) and "30" in str(h8))


def is_slim_layout(ws) -> bool:
    h10 = ws.cell(HDR_ROW, COL_HOLD).value
    h11 = ws.cell(HDR_ROW, 11).value
    return bool(h10 and "Hold" in str(h10) and not h11)


def ensure_np_columns(ws) -> None:
    if not has_np_columns(ws):
        ws.insert_cols(COL_NP30, 2)


def trim_extra_columns(ws) -> None:
    if is_slim_layout(ws):
        while ws.max_column > LAST_COL:
            ws.delete_cols(LAST_COL + 1, 1)
        return
    # Standard wide layout after NP insert: drop purch / personality / comm cols
    for col in (14, 13, 11, 10):
        if ws.max_column >= col:
            ws.delete_cols(col, 1)
    while ws.max_column > LAST_COL:
        ws.delete_cols(LAST_COL + 1, 1)


def compute_top_state(ws) -> str:
    states = [ws.cell(r, COL_STATE).value for r in range(DATA_START, DATA_END + 1)]
    counts = Counter(s for s in states if s)
    if not counts:
        return "—"
    name, n = counts.most_common(1)[0]
    return f"{name} · {n} of 50 ({n / 50:.0%})"


def clear_kpi_rows(ws) -> None:
    for row in (3, 4, 5):
        for col in range(1, 20):
            ws.cell(row, col).value = None
            ws.cell(row, col).fill = PatternFill(fill_type=None)


def style_kpi_block(ws, top_state_label: str) -> None:
    fill(ws["A1"], "VIP Event — Top 50 Players", FONT_TITLE, PatternFill("solid", fgColor=NAVY))
    ws.row_dimensions[1].height = 34

    for col in range(1, LAST_COL + 1):
        fill(ws.cell(3, col), font=FONT_SECTION, fill_=PatternFill("solid", fgColor=METRIC_BG))
    ws["A3"] = "  COHORT AVERAGES"
    ws.row_dimensions[3].height = 16

    metrics = [
        ("=AVERAGE(D9:D58)", "0"),
        ("=AVERAGE(G9:G58)", MONEY_FMT),
        ("=AVERAGE(H9:H58)", MONEY_FMT),
        ("=AVERAGE(I9:I58)", MONEY_FMT),
        ("=AVERAGE(J9:J58)", PCT_FMT),
        (top_state_label, None),
    ]
    labels = [
        "AVG AGE",
        "AVG NET PURCHASE\nLIFETIME",
        "AVG NET PURCHASE\n30D",
        "AVG NET PURCHASE\n60D",
        "AVG HOLD %",
        "TOP STATE",
    ]

    white = PatternFill("solid", fgColor=WHITE)
    state_font = Font(name="Calibri", size=12, bold=True, color="FF1A1A2E")
    for col_idx, (formula, fmt) in zip(KPI_COLS, metrics):
        c = ws.cell(4, col_idx)
        font = state_font if col_idx == 16 else FONT_KPI_VALUE
        fill(c, formula, font, white, KPI_ALIGN)
        if fmt:
            c.number_format = fmt

    for col_idx, text in zip(KPI_COLS, labels):
        fill(ws.cell(5, col_idx), text, FONT_KPI_LABEL, white, KPI_LABEL_ALIGN)

    ws.row_dimensions[4].height = 32
    ws.row_dimensions[5].height = 26

    for col_idx in KPI_COLS:
        ws.column_dimensions[get_column_letter(col_idx)].width = 18 if col_idx == 16 else 13

    ws.row_dimensions[6].height = 6
    ws.cell(7, 1).value = None


def style_table_headers(ws) -> None:
    hdr_fill = PatternFill("solid", fgColor=HEADER_FILL)
    for col, label in TABLE_HEADERS.items():
        fill(
            ws.cell(HDR_ROW, col),
            label,
            FONT_TABLE_HDR,
            hdr_fill,
            Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=CELL_BORDER,
        )
    ws.row_dimensions[HDR_ROW].height = 36


def style_data_rows(ws, metrics: dict[int, dict[str, float]]) -> None:
    for row in range(DATA_START, DATA_END + 1):
        alt = row % 2 == 1
        row_fill = PatternFill("solid", fgColor=ALT_FILL if alt else WHITE)
        aid = int(ws.cell(row, COL_AID).value)
        m = metrics.get(aid, {"net_purchase_30d": 0.0, "net_purchase_60d": 0.0})

        for col in range(1, LAST_COL + 1):
            c = ws.cell(row, col)
            c.fill = row_fill
            c.border = CELL_BORDER
            c.alignment = Alignment(
                horizontal="right" if col >= COL_AGE and col != COL_STATE else "left",
                vertical="center",
            )

        ws.cell(row, COL_RANK).font = FONT_RANK
        ws.cell(row, COL_RANK).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, COL_NAME).font = FONT_NAME
        ws.cell(row, COL_AID).font = FONT_BODY
        ws.cell(row, COL_AGENT).font = FONT_BODY

        for col in (COL_LT, COL_NP30, COL_NP60):
            c = ws.cell(row, col)
            c.font = FONT_MONEY
            c.number_format = MONEY_FMT

        fill(
            ws.cell(row, COL_NP30),
            round(m["net_purchase_30d"], 2),
            FONT_MONEY,
            row_fill,
            Alignment(horizontal="right", vertical="center"),
            MONEY_FMT,
            CELL_BORDER,
        )
        fill(
            ws.cell(row, COL_NP60),
            round(m["net_purchase_60d"], 2),
            FONT_MONEY,
            row_fill,
            Alignment(horizontal="right", vertical="center"),
            MONEY_FMT,
            CELL_BORDER,
        )

        hold = ws.cell(row, COL_HOLD)
        if hold.value is not None:
            hold.number_format = PCT_FMT

    widths = {1: 6, 2: 12, 3: 22, 4: 6, 5: 14, 6: 12, 7: 16, 8: 14, 9: 14, 10: 10}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def style_total_row(ws) -> None:
    total_fill = PatternFill("solid", fgColor=TOTAL_FILL)
    for col in range(1, LAST_COL + 1):
        c = ws.cell(TOTAL_ROW, col)
        c.fill = total_fill
        c.border = CELL_BORDER
        c.font = FONT_TOTAL

    fill(ws.cell(TOTAL_ROW, COL_RANK), "TOTAL", FONT_TOTAL, total_fill, Alignment(horizontal="left"))
    fill(
        ws.cell(TOTAL_ROW, COL_NAME),
        "50 players",
        FONT_TOTAL,
        total_fill,
        Alignment(horizontal="left"),
    )

    ws.cell(TOTAL_ROW, COL_AGE).value = "=SUM(D9:D58)"
    ws.cell(TOTAL_ROW, COL_AGE).number_format = "0"
    ws.cell(TOTAL_ROW, COL_AGE).alignment = Alignment(horizontal="right")

    for col, formula in (
        (COL_LT, "=SUM(G9:G58)"),
        (COL_NP30, "=SUM(H9:H58)"),
        (COL_NP60, "=SUM(I9:I58)"),
    ):
        c = ws.cell(TOTAL_ROW, col)
        c.value = formula
        c.number_format = MONEY_FMT
        c.alignment = Alignment(horizontal="right")

    # Hold % — sum is not meaningful; show cohort average in totals row
    c = ws.cell(TOTAL_ROW, COL_HOLD)
    c.value = "=AVERAGE(J9:J58)"
    c.number_format = PCT_FMT
    c.alignment = Alignment(horizontal="right")

    ws.row_dimensions[TOTAL_ROW].height = 22


def remove_extra_sections(ws) -> None:
    for row in range(TOTAL_ROW + 1, ws.max_row + 1):
        for col in range(1, 25):
            cell = ws.cell(row, col)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)


def layout_workbook(ws, metrics: dict[int, dict[str, float]]) -> None:
    ensure_np_columns(ws)
    trim_extra_columns(ws)
    remove_extra_sections(ws)
    clear_kpi_rows(ws)
    top_state = compute_top_state(ws)
    style_kpi_block(ws, top_state)
    style_table_headers(ws)
    style_data_rows(ws, metrics)
    style_total_row(ws)


def enrich(source: Path, report_date: date, output: Path | None = None) -> Path:
    work_path = copy_workbook_source(source)
    wb = load_workbook(work_path)
    ws = wb[SHEET_NAME]

    aids = [int(ws.cell(r, COL_AID).value) for r in range(DATA_START, DATA_END + 1)]
    metrics = fetch_net_purchases(aids, report_date)
    layout_workbook(ws, metrics)

    out = output or source
    try:
        wb.save(out)
    except PermissionError:
        fallback = DATA_DIR / "VIP Event - Top 50 Players - enriched.xlsx"
        wb.save(fallback)
        print(f"Could not write target file (open in Excel?). Saved: {fallback}")
        return fallback

    wb.save(work_path)
    print(f"Wrote: {out}")
    print(f"Workspace copy: {work_path}")
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="Enrich VIP Event Top 50 xlsx")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument(
        "--date",
        type=lambda s: date.fromisoformat(s),
        default=date.today() - timedelta(days=1),
    )
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()
    enrich(args.source, args.date, args.output)


if __name__ == "__main__":
    main()
