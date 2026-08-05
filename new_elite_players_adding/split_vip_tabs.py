import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

INPUT_FILE  = r"c:\Users\Owner\Downloads\Elite\new_elite_players_adding\VIP Potential.csv"
OUTPUT_FILE = r"c:\Users\Owner\Downloads\Elite\new_elite_players_adding\VIP_Potential_Split.xlsx"

RANKING_METRICS = [
    "Avg. LT_net_purchases_ByReq",
    "Avg. LT_purchased",
    "LT Hold",
    "Previous 30d Purchased *",
]

# Tab names — order determines metric pool priority (metric 0→1→2→3)
AM_CONFIG = [
    {"key": "Coral",   "metric": RANKING_METRICS[0]},
    {"key": "Lee",     "metric": RANKING_METRICS[1]},
    {"key": "Rachel",  "metric": RANKING_METRICS[2]},
    {"key": "Gabriel", "metric": RANKING_METRICS[3]},
]

# ARGB hex strings (FF + RGB)
PALETTE = {
    "Coral":   {"header": "FF0070C0", "light": "FFD0E8F8", "accent": "FF00B0F0", "tab": "00B0F0"},
    "Lee":     {"header": "FF008000", "light": "FFD0F0D0", "accent": "FF00B050", "tab": "00B050"},
    "Rachel":  {"header": "FFB83200", "light": "FFFFE0CC", "accent": "FFFF6600", "tab": "FF6600"},
    "Gabriel": {"header": "FF6600AA", "light": "FFEED5F5", "accent": "FF9900CC", "tab": "9900CC"},
    "Master":  {"header": "FF404040", "light": "FFF2F2F2", "accent": "FF808080", "tab": "808080"},
}

# Fixed column order shown on every AM tab
ID_COLS       = ["account_id", "first_name", "last_name", "email"]
DISPLAY_COLS  = ID_COLS + [
    "Avg. LT_net_purchases_ByReq",
    "Avg. LT_purchased",
    "LT Hold",
    "Previous 30d Purchased *",
]

# Columns whose raw values look like percentages (keep as 0–100 floats, format as %)
PCT_HINTS = {"LT Hold", "LT Ent. Rate", "Entertainment Rate", "Hold %", "Margin %"}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_numeric(series: pd.Series) -> pd.Series:
    """Strip currency/percent formatting and return float."""
    return (
        series.astype(str)
        .str.replace(r"[\$,\s%]", "", regex=True)
        .pipe(pd.to_numeric, errors="coerce")
    )


def detect_numeric_cols(df: pd.DataFrame, threshold: float = 0.80) -> set[str]:
    """Return column names where ≥ threshold fraction of rows parse as numbers."""
    numeric_cols = set()
    for col in df.columns:
        parsed = clean_numeric(df[col])
        frac = parsed.notna().sum() / max(len(df), 1)
        if frac >= threshold:
            numeric_cols.add(col)
    return numeric_cols


def fill(argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=argb)


def write_sheet(
    ws,
    df: pd.DataFrame,
    palette: dict,
    numeric_cols: set[str],
    totals_col: str | None = "Net Purchases",
) -> None:
    """Write a fully-styled sheet. Numeric columns are written as floats with formats."""

    cols = list(df.columns)

    # ---------- styles ----------
    hdr_fill  = fill(palette["header"])
    light_fill = fill(palette["light"])
    alt_fill   = fill("FFF7F7F7")
    total_fill = fill(palette["accent"])

    hdr_font   = Font(name="Calibri", bold=True,  color="FFFFFF", size=11)
    body_font  = Font(name="Calibri",              size=10)
    total_font = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    thin = Border(bottom=Side(style="thin", color="E0E0E0"))

    # ---------- number formats ----------
    def excel_format(col_name: str) -> str | None:
        if col_name not in numeric_cols:
            return None
        if col_name in PCT_HINTS:
            return '0.0"%"'
        return "#,##0"

    # ---------- header row ----------
    for c_idx, name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center

    # ---------- data rows ----------
    data_rows = len(df)
    for r_idx, row_vals in enumerate(df.itertuples(index=False), 2):
        is_alt = (r_idx % 2 == 0)
        for c_idx, (col_name, raw_val) in enumerate(zip(cols, row_vals), 1):
            fmt = excel_format(col_name)
            if fmt and raw_val is not None:
                try:
                    val = float(str(raw_val).replace("$", "").replace(",", "").replace("%", "").replace(" ", ""))
                except (ValueError, TypeError):
                    val = raw_val
            else:
                val = raw_val

            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font   = body_font
            cell.border = thin
            if is_alt:
                cell.fill = alt_fill

            if fmt:
                cell.number_format = fmt
                cell.alignment = right
            elif col_name in ("#", "account_id"):
                cell.alignment = center
            else:
                cell.alignment = left

    # ---------- totals row ----------
    if totals_col and totals_col in cols:
        t_row = data_rows + 2
        tc_idx = cols.index(totals_col) + 1

        # Label in first non-# column
        label_idx = 2 if cols[0] == "#" else 1
        label_cell = ws.cell(row=t_row, column=label_idx, value="TOTAL Avg. LT Net Purchases")
        label_cell.fill = total_fill
        label_cell.font = total_font
        label_cell.alignment = left

        col_letter = get_column_letter(tc_idx)
        total_cell = ws.cell(
            row=t_row, column=tc_idx,
            value=f"=SUM({col_letter}2:{col_letter}{data_rows + 1})",
        )
        total_cell.fill = total_fill
        total_cell.font = total_font
        total_cell.number_format = "#,##0"
        total_cell.alignment = right

        # Fill remaining cells in totals row with accent color
        for c_idx in range(1, len(cols) + 1):
            cell = ws.cell(row=t_row, column=c_idx)
            cell.fill = total_fill

    # ---------- freeze + row heights ----------
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 18

    # ---------- auto-fit column widths ----------
    for c_idx, col_name in enumerate(cols, 1):
        col_letter = get_column_letter(c_idx)
        sample = range(1, min(ws.max_row + 1, 60))
        max_len = max(
            (len(str(ws.cell(row=r, column=c_idx).value or "")) for r in sample),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # Tab color
    ws.sheet_properties.tabColor = palette["tab"]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    # 1. Read (UTF-16 tab-separated)
    df = pd.read_csv(INPUT_FILE, sep="\t", dtype=str, encoding="utf-16")
    # Strip stray whitespace from column names
    df.columns = df.columns.str.strip()
    print(f"Loaded {len(df)} rows, {len(df.columns)} columns.")

    # 2. Detect numeric columns
    numeric_cols = detect_numeric_cols(df)
    print(f"Numeric columns detected: {len(numeric_cols)}")

    # 3. Build sort keys for metric selection and draft
    sort_keys: dict[str, str] = {}
    for metric in RANKING_METRICS:
        key = f"__sort_{metric}"
        df[key] = clean_numeric(df[metric])
        sort_keys[metric] = key

    # 4. Pool selection — greedy priority (metric 1 → 2 → 3 → 4), no repeats
    pool = df.copy()
    pool_200_ids: list[str] = []
    for cfg in AM_CONFIG:
        sk    = sort_keys[cfg["metric"]]
        top50 = pool.sort_values(sk, ascending=False).head(50)
        pool_200_ids.extend(top50["account_id"].tolist())
        pool = pool[~pool["account_id"].isin(top50["account_id"])]

    candidates = df[df["account_id"].isin(pool_200_ids)].copy()
    print(f"Pool of 200 candidates selected.")

    # 5. Draft-balance by Avg. LT_net_purchases_ByReq
    #    Sort all 200 by that metric desc, assign each player to the tab with
    #    the lowest running total — ensures balanced sums across all tabs.
    BALANCE_METRIC = "Avg. LT_net_purchases_ByReq"
    candidates = candidates.sort_values(sort_keys[BALANCE_METRIC], ascending=False)

    tab_totals: dict[str, float] = {cfg["key"]: 0.0 for cfg in AM_CONFIG}
    tab_rows:   dict[str, list]  = {cfg["key"]: []   for cfg in AM_CONFIG}

    for _, row in candidates.iterrows():
        # Pick the tab with lowest running total; skip any tab already at 50
        eligible = [k for k in tab_totals if len(tab_rows[k]) < 50]
        target = min(eligible, key=lambda k: (tab_totals[k], len(tab_rows[k])))
        tab_rows[target].append(row)
        bal_val = row[sort_keys[BALANCE_METRIC]]
        tab_totals[target] += float(bal_val) if pd.notna(bal_val) else 0.0

    print(f"Draft results (Avg. LT_net_purchases_ByReq totals per tab):")
    for am_key, total in tab_totals.items():
        print(f"  {am_key}: {total:,.0f}  ({len(tab_rows[am_key])} players)")

    # 6. Build per-tab DataFrames (sorted by Avg. LT_net_purchases_ByReq desc)
    tab_frames: dict[str, pd.DataFrame] = {}
    assigned:   dict[str, str]          = {}
    for cfg in AM_CONFIG:
        am_key = cfg["key"]
        tab_df = pd.DataFrame(tab_rows[am_key])
        tab_df = tab_df.sort_values(sort_keys["Avg. LT_net_purchases_ByReq"], ascending=False)
        tab_frames[am_key] = tab_df
        for aid in tab_df["account_id"]:
            assigned[aid] = am_key

    # 7. Tag master df
    df["Assigned To"] = df["account_id"].map(assigned).fillna("")

    # 8. Drop all sort-key helper columns
    all_sort_keys = list(sort_keys.values())
    df.drop(columns=all_sort_keys, inplace=True, errors="ignore")
    for am_key in tab_frames:
        tab_frames[am_key].drop(columns=all_sort_keys, inplace=True, errors="ignore")

    # 9. Build workbook
    wb = Workbook()
    wb.remove(wb.active)

    # --- AM tabs (Coral / Lee / Rachel / Gabriel) ---
    for cfg in AM_CONFIG:
        am_key  = cfg["key"]
        palette = PALETTE[am_key]

        am_df = tab_frames[am_key].copy()

        # Fixed column order — same on all tabs
        ordered = [c for c in DISPLAY_COLS if c in am_df.columns]
        # Any extra columns not in DISPLAY_COLS are omitted from AM tabs
        am_df = am_df[ordered].reset_index(drop=True)
        am_df.insert(0, "#", range(1, len(am_df) + 1))

        ws = wb.create_sheet(title=am_key)
        write_sheet(ws, am_df, palette, numeric_cols, totals_col="Avg. LT_net_purchases_ByReq")
        print(f"  Sheet '{am_key}': {len(am_df)} players written.")

    # --- Master tab ---
    front_cols = ["account_id", "first_name", "last_name", "email", "Assigned To"] + RANKING_METRICS
    rest_cols  = [c for c in df.columns if c not in front_cols]
    master_df  = df[front_cols + rest_cols].copy()
    # Sort: assigned (by name) first, then unassigned
    master_df["__sort_am"] = master_df["Assigned To"].replace("", "ZZ")
    master_df = master_df.sort_values("__sort_am").drop(columns=["__sort_am"])
    master_df.insert(0, "#", range(1, len(master_df) + 1))

    ws_master = wb.create_sheet(title="Master")
    write_sheet(ws_master, master_df, PALETTE["Master"], numeric_cols, totals_col="Avg. LT_net_purchases_ByReq")

    # Move Master to first position
    wb.move_sheet("Master", offset=-(len(wb.sheetnames) - 1))

    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
