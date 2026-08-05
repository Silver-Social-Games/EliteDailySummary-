from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DISPLAY_COLS = [
    "account_id",
    "first_name",
    "last_name",
    "email",
    "Avg. LT_net_purchases_ByReq",
    "Avg. LT_purchased",
    "LT Hold",
    "Previous 30d Purchased *",
]
PERCENT_COLS = {
    "LT Hold",
    "Hold %",
    "Margin %",
    "Entertainment Rate",
    "LT Ent. Rate",
}
PALETTES = [
    {"header": "FF0070C0", "tab": "00B0F0"},
    {"header": "FF6600AA", "tab": "9900CC"},
    {"header": "FF008000", "tab": "00B050"},
    {"header": "FFB83200", "tab": "FF6600"},
]
MASTER_PALETTE = {"header": "FF404040", "tab": "808080"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Assign New Elite players to exact Agent quotas."
    )
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--quota",
        action="append",
        required=True,
        metavar="AGENT=COUNT",
        help="Repeat for each Agent.",
    )
    parser.add_argument("--base", type=Path)
    parser.add_argument("--balance-column", default="Net Purchases")
    parser.add_argument("--sheet-suffix", default="New")
    parser.add_argument("--master-sheet", default="New Additions Master")
    parser.add_argument(
        "--show-balance-column",
        action="store_true",
        help="Include the balancing column in generated sheets.",
    )
    parser.add_argument(
        "--remove-balance-from-all-sheets",
        action="store_true",
        help="Remove the balancing column from inherited sheets too.",
    )
    parser.add_argument(
        "--force-assign",
        action="append",
        default=[],
        metavar="AGENT=AID",
        help="Force an AID onto an Agent before balanced assignment. Repeatable.",
    )
    return parser.parse_args()


def parse_quotas(values: list[str]) -> dict[str, int]:
    quotas: dict[str, int] = {}
    for value in values:
        if "=" not in value:
            raise ValueError(f"Invalid quota '{value}'; expected AGENT=COUNT.")
        agent, raw_count = value.rsplit("=", 1)
        agent = agent.strip()
        count = int(raw_count)
        if not agent or count < 0:
            raise ValueError(f"Invalid quota '{value}'.")
        if agent in quotas:
            raise ValueError(f"Duplicate quota for '{agent}'.")
        quotas[agent] = count
    return quotas


def parse_force_assigns(values: list[str], quotas: dict[str, int]) -> dict[str, str]:
    forced: dict[str, str] = {}
    for value in values:
        if "=" not in value:
            raise ValueError(f"Invalid force-assign '{value}'; expected AGENT=AID.")
        agent, account_id = value.split("=", 1)
        agent = agent.strip()
        account_id = account_id.strip()
        if agent not in quotas:
            raise ValueError(f"Force-assign agent '{agent}' is not in quotas.")
        if not account_id:
            raise ValueError(f"Invalid force-assign '{value}'.")
        if account_id in forced:
            raise ValueError(f"AID {account_id} forced more than once.")
        forced[account_id] = agent
    return forced


def clean_numeric(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.replace(r"[\$,\s%]", "", regex=True)
        .pipe(pd.to_numeric, errors="coerce")
    )


def detect_numeric_cols(df: pd.DataFrame, threshold: float = 0.80) -> set[str]:
    return {
        col
        for col in df.columns
        if clean_numeric(df[col]).notna().sum() / max(len(df), 1) >= threshold
    }


def assign_players(
    df: pd.DataFrame,
    quotas: dict[str, int],
    balance_column: str,
    force_assigns: dict[str, str] | None = None,
) -> tuple[dict[str, pd.DataFrame], dict[str, float]]:
    force_assigns = force_assigns or {}
    work = df.copy()
    work["__balance"] = clean_numeric(work[balance_column]).fillna(0)
    work["account_id"] = work["account_id"].astype(str)
    rows: dict[str, list[pd.Series]] = {agent: [] for agent in quotas}
    totals = {agent: 0.0 for agent in quotas}

    missing = sorted(set(force_assigns) - set(work["account_id"]))
    if missing:
        raise ValueError(f"Forced AIDs missing from CSV: {', '.join(missing)}")

    for account_id, agent in force_assigns.items():
        if len(rows[agent]) >= quotas[agent]:
            raise ValueError(
                f"Force-assign to {agent} exceeds quota {quotas[agent]}."
            )
        player = work.loc[work["account_id"] == account_id].iloc[0]
        rows[agent].append(player)
        totals[agent] += float(player["__balance"])

    remaining = work[~work["account_id"].isin(force_assigns)].sort_values(
        "__balance", ascending=False
    )
    for _, player in remaining.iterrows():
        eligible = [
            agent for agent, quota in quotas.items() if len(rows[agent]) < quota
        ]
        target = min(eligible, key=lambda agent: (totals[agent], len(rows[agent])))
        rows[target].append(player)
        totals[target] += float(player["__balance"])

    frames: dict[str, pd.DataFrame] = {}
    for agent, agent_rows in rows.items():
        frame = pd.DataFrame(agent_rows).sort_values("__balance", ascending=False)
        frames[agent] = frame.drop(columns=["__balance"])
    return frames, totals


def remove_column_from_workbook(wb: Workbook, column_name: str) -> None:
    for ws in wb.worksheets:
        headers = [cell.value for cell in ws[1]]
        if column_name in headers:
            ws.delete_cols(headers.index(column_name) + 1)
        for row_idx in range(ws.max_row, 1, -1):
            values = [cell.value for cell in ws[row_idx]]
            if any(value == f"TOTAL {column_name}" for value in values):
                ws.delete_rows(row_idx)


def add_or_replace_sheet(wb: Workbook, title: str):
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title=title)


def write_sheet(
    ws,
    df: pd.DataFrame,
    palette: dict[str, str],
    numeric_cols: set[str],
) -> None:
    columns = list(df.columns)
    header_fill = PatternFill("solid", fgColor=palette["header"])
    alternate_fill = PatternFill("solid", fgColor="FFF7F7F7")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Calibri", size=10)
    border = Border(bottom=Side(style="thin", color="E0E0E0"))
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    for col_idx, name in enumerate(columns, 1):
        cell = ws.cell(1, col_idx, name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row_idx, values in enumerate(df.itertuples(index=False), 2):
        for col_idx, (column, raw_value) in enumerate(zip(columns, values), 1):
            is_numeric = column in numeric_cols
            value = raw_value
            if is_numeric and raw_value is not None:
                parsed = clean_numeric(pd.Series([raw_value])).iloc[0]
                value = float(parsed) if pd.notna(parsed) else raw_value

            cell = ws.cell(row_idx, col_idx, value)
            cell.font = body_font
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alternate_fill
            if is_numeric:
                cell.number_format = (
                    '0.0"%"' if column in PERCENT_COLS else "#,##0"
                )
                cell.alignment = right
            elif column in {"#", "account_id"}:
                cell.alignment = center
            else:
                cell.alignment = left

    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = palette["tab"]
    ws.row_dimensions[1].height = 22
    for col_idx in range(1, len(columns) + 1):
        width = max(
            len(str(ws.cell(row_idx, col_idx).value or ""))
            for row_idx in range(1, min(ws.max_row, 80) + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 3, 42)


def load_output_workbook(base: Path | None) -> Workbook:
    if base:
        if not base.exists():
            raise FileNotFoundError(f"Base workbook not found: {base}")
        return load_workbook(base)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def main() -> None:
    args = parse_args()
    quotas = parse_quotas(args.quota)
    force_assigns = parse_force_assigns(args.force_assign, quotas)
    df = pd.read_csv(args.input, sep="\t", dtype=str, encoding="utf-16")
    df.columns = df.columns.str.strip()

    required = set(DISPLAY_COLS) | {args.balance_column}
    missing = sorted(required - set(df.columns))
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    if len(df) != sum(quotas.values()):
        raise ValueError(
            f"Quota total is {sum(quotas.values())}, but CSV has {len(df)} rows."
        )
    if df["account_id"].duplicated().any():
        raise ValueError("Source CSV contains duplicate account_id values.")

    numeric_cols = detect_numeric_cols(df)
    frames, totals = assign_players(
        df, quotas, args.balance_column, force_assigns=force_assigns
    )
    wb = load_output_workbook(args.base)
    if args.remove_balance_from_all_sheets:
        remove_column_from_workbook(wb, args.balance_column)

    display_cols = DISPLAY_COLS.copy()
    if args.show_balance_column and args.balance_column not in display_cols:
        display_cols.insert(4, args.balance_column)

    assigned_to: dict[str, str] = {}
    for index, (agent, quota) in enumerate(quotas.items()):
        frame = frames[agent]
        assigned_to.update(
            {str(account_id): agent for account_id in frame["account_id"]}
        )
        output = frame[display_cols].reset_index(drop=True)
        output.insert(0, "#", range(1, len(output) + 1))
        title = f"{agent} {args.sheet_suffix}".strip()
        ws = add_or_replace_sheet(wb, title)
        write_sheet(ws, output, PALETTES[index % len(PALETTES)], numeric_cols)
        print(
            f"{agent}: {len(output)}/{quota} players; "
            f"{args.balance_column}={totals[agent]:,.0f}"
        )

    master = df.copy()
    master["Assigned To"] = (
        master["account_id"].astype(str).map(assigned_to).fillna("")
    )
    master_cols = [
        "account_id",
        "first_name",
        "last_name",
        "email",
        "Assigned To",
        *[col for col in display_cols if col not in DISPLAY_COLS[:4]],
    ]
    master = master[master_cols].sort_values(["Assigned To", "account_id"])
    master.insert(0, "#", range(1, len(master) + 1))
    ws = add_or_replace_sheet(wb, args.master_sheet)
    write_sheet(ws, master, MASTER_PALETTE, numeric_cols)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    assigned = sum(len(frame) for frame in frames.values())
    unique = len(assigned_to)
    if assigned != len(df) or unique != len(df):
        raise RuntimeError(
            f"Verification failed: assigned={assigned}, unique={unique}, "
            f"source={len(df)}."
        )
    print(f"Verified: {assigned} assigned, {unique} unique AIDs.")
    print(f"Saved: {args.output}")


if __name__ == "__main__":
    main()
