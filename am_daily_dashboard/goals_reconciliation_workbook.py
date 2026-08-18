"""Fill-in workbook for reconciling the Elite Goals board against an AM's own
table.

Built because the interactive canvas would not open on this machine. Same idea,
delivered as Excel: one sheet per AM, an empty **Your value** column on the left,
the board's figure beside it, and live gap formulas so a mismatch is visible as
soon as a number is typed.

Layout is deliberately two-layered. Layer 1 is the nine raw figures read straight
out of BigQuery; layer 2 is the seven scored KPIs, which are pure arithmetic on
those nine. So a gap in layer 1 is a root cause, while a gap that appears only in
layer 2 points at the board's formula instead. Fill Tagged accounts first: one
wrong book size moves all seven KPIs at once.

Run:  python am_daily_dashboard/goals_reconciliation_workbook.py
Reads the Goals blocks from the newest AM Brief JSON export, so it never
hard-codes numbers that could drift out of date.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PACKAGE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(PACKAGE_DIR.parent))

from elite_lib.export_paths import mirror_to_cursor  # noqa: E402

EXPORTS = PACKAGE_DIR / "exports"
OUT_NAME = "Elite Goals reconciliation.xlsx"

INPUT_FILL = PatternFill("solid", fgColor="FFF6D9")
HEAD_FILL = PatternFill("solid", fgColor="EDEDED")
SECTION_FILL = PatternFill("solid", fgColor="E4ECF7")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

USD = '"$"#,##0'
CNT = "#,##0"
PCT = '0.0"%"'

# (label, payload key or derived marker, number format, where it comes from)
RAW_ROWS = [
    ("Tagged accounts (whole book)", "portfolioSizeAll", CNT,
     "dbt_aninditac.elite + tag_agent_1, snapshot pinned to the report date"),
    ("Locked accounts", "portfolioLocked", CNT, "uam_accounts.locked"),
    ("Eligible portfolio (% Active denominator)", "portfolioSize", CNT,
     "unlocked, plus locked accounts that still bought this month"),
    ("MTD Purchase", "mtdPurchase", USD, "sum of successful payment orders"),
    ("MTD Net Purchase", "mtdNetPurchase", USD,
     "purchase - requested redeem (net of cancelled) - chargeback - refund"),
    ("Distinct purchasers in month", "kpi:monthly_purchasers", CNT,
     "distinct AID with a successful order in the month"),
    ("Reactivated AIDs", "kpi:reactivations", CNT,
     "purchase after a gap of 20+ days, counted once per AID"),
    ("Active players (% Active numerator)", "derived:active", CNT,
     "last successful purchase within 30 days of the as-of date"),
    ("Upgraded to Elite in month", "kpi:upgrades", CNT,
     "first Elite tag snapshot in month, not Elite before it"),
]

KPI_ROWS = [
    ("Daily Avg Purchase", "daily_avg_purchase", USD,
     "MTD Purchase / elapsed days"),
    ("Daily Avg Net Purchase", "daily_avg_net_purchase", USD,
     "MTD Net Purchase / elapsed days"),
    ("Monthly Purchasers", "monthly_purchasers", CNT,
     "same as the raw distinct purchaser count"),
    ("ARPPU", "arppu", USD, "MTD Purchase / distinct purchasers"),
    ("# Reactivation", "reactivations", CNT,
     "same as the raw reactivated count"),
    ("Upgrade to Elite", "upgrades", CNT, "same as the raw upgrade count"),
    ("% Active from portfolio", "pct_active", PCT,
     "active players / eligible portfolio"),
]

QUESTIONS = [
    ("Which report are your correct numbers from?",
     "Tableau daily-agg, the Goals sheet, or a manual export each define the "
     "book differently. Reactivation already matches Tableau at a 20-day gap."),
    ("What exact date range does it cover?",
     "The board is the 1st to the as-of date inclusive. A range ending a day "
     "later shifts every daily average and every count."),
    ("Does your book count the whole tagged roster or only part of it?",
     "Highest-leverage single number. If Tagged accounts differs, every KPI "
     "inherits it and there is one root cause instead of seven."),
    ("For revenue, is your figure gross of refunds and chargebacks?",
     "Board MTD Purchase is gross successful orders; Net subtracts requested "
     "redeem, chargeback and refund."),
    ("Is Upgrade to Elite dated by tag date or by first purchase?",
     "The board uses the first Elite tag snapshot in the month. Dating by FTP "
     "or by assignment date diverges."),
]


def newest_export() -> Path:
    files = sorted(EXPORTS.glob("*_elite_am_brief.json"))
    if not files:
        raise SystemExit(
            f"No AM Brief JSON in {EXPORTS}. Run "
            "generate_am_daily_dashboard.py first."
        )
    return files[-1]


def kpi_map(goals: dict) -> dict[str, dict]:
    return {k.get("key"): k for k in goals.get("kpis", []) if k.get("key")}


def board_value(goals: dict, key: str) -> float | None:
    kpis = kpi_map(goals)
    if key.startswith("kpi:"):
        kpi = kpis.get(key[4:])
        return None if kpi is None else kpi.get("actual")
    if key == "derived:active":
        pct = kpis.get("pct_active", {}).get("actual")
        size = goals.get("portfolioSize")
        if pct is None or not size:
            return None
        return round(pct / 100.0 * size)
    return goals.get(key)


def write_section(ws, row: int, title: str, note: str) -> int:
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    for col in range(2, 7):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    row += 1
    cell = ws.cell(row=row, column=1, value=note)
    cell.font = Font(italic=True, size=9, color="595959")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 28
    row += 1

    headers = ["Your value", "Metric", "Board", "Gap", "Gap %", "How the board gets it"]
    for col, head in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=head)
        c.font = Font(bold=True)
        c.fill = HEAD_FILL
        c.border = BOX
    return row + 1


def write_rows(ws, row: int, rows: list[tuple], goals: dict, kpi_mode: bool) -> int:
    for label, key, fmt, source in rows:
        value = (
            kpi_map(goals).get(key, {}).get("actual")
            if kpi_mode
            else board_value(goals, key)
        )
        yours = ws.cell(row=row, column=1)
        yours.fill = INPUT_FILL
        yours.border = BOX
        yours.number_format = fmt

        ws.cell(row=row, column=2, value=label).border = BOX

        board_cell = ws.cell(row=row, column=3, value=value)
        board_cell.number_format = fmt
        board_cell.border = BOX

        gap = ws.cell(row=row, column=4, value=f"=IF(A{row}=\"\",\"\",A{row}-C{row})")
        gap.number_format = fmt
        gap.border = BOX

        gap_pct = ws.cell(
            row=row,
            column=5,
            value=f'=IF(OR(A{row}="",C{row}=0),"",(A{row}-C{row})/C{row})',
        )
        gap_pct.number_format = "+0.0%;-0.0%"
        gap_pct.border = BOX

        src = ws.cell(row=row, column=6, value=source)
        src.font = Font(size=9, color="595959")
        src.alignment = Alignment(wrap_text=True, vertical="top")
        src.border = BOX
        row += 1
    return row + 1


def build_am_sheet(wb: Workbook, name: str, goals: dict, as_of: str) -> None:
    ws = wb.create_sheet(name)
    ws["A1"] = f"{name} — Elite Goals reconciliation"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        f"{goals.get('monthLabel', '')} as of {as_of} · "
        f"{goals.get('elapsedDays', 0)} of {goals.get('daysInMonth', 0)} days · "
        f"book snapshot {goals.get('bookSnapshotDate') or 'n/a'} · "
        "type your figures in the shaded column A"
    )
    ws["A2"].font = Font(size=10, color="595959")

    row = write_section(
        ws,
        4,
        "Layer 1 — raw inputs from BigQuery",
        "A gap here is a root cause. Fill Tagged accounts first: one wrong book "
        "size moves all seven KPIs below at once.",
    )
    row = write_rows(ws, row, RAW_ROWS, goals, kpi_mode=False)

    row = write_section(
        ws,
        row,
        "Layer 2 — the seven scored KPIs",
        "These are pure arithmetic on layer 1. If every raw input above matches "
        "but a KPI here does not, the board's formula is at fault.",
    )
    write_rows(ws, row, KPI_ROWS, goals, kpi_mode=True)

    widths = [14, 42, 16, 14, 10, 52]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


def build_questions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Questions", 0)
    ws["A1"] = "Answer these and each gap can be traced to a cause"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Type answers in column B. One line each is enough."
    ws["A2"].font = Font(size=10, color="595959")

    for col, head in enumerate(["Question", "Your answer", "Why it matters"], start=1):
        c = ws.cell(row=4, column=col, value=head)
        c.font = Font(bold=True)
        c.fill = HEAD_FILL
        c.border = BOX

    row = 5
    for question, why in QUESTIONS:
        q = ws.cell(row=row, column=1, value=question)
        q.alignment = Alignment(wrap_text=True, vertical="top")
        q.border = BOX
        ans = ws.cell(row=row, column=2)
        ans.fill = INPUT_FILL
        ans.border = BOX
        ans.alignment = Alignment(wrap_text=True, vertical="top")
        w = ws.cell(row=row, column=3, value=why)
        w.font = Font(size=9, color="595959")
        w.alignment = Alignment(wrap_text=True, vertical="top")
        w.border = BOX
        ws.row_dimensions[row].height = 46
        row += 1

    for letter, width in (("A", 52), ("B", 44), ("C", 62)):
        ws.column_dimensions[letter].width = width


def has_typed_answers(path: Path) -> bool:
    """True if a previous copy already holds the user's own figures.

    Regenerating silently destroyed a filled-in copy once (2026-08-18). Column A
    is the only place the user types on an AM sheet, and generated sheets leave it
    empty, so any value there means real work would be lost.
    """
    if not path.is_file():
        return False
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path)
    except Exception:
        # Unreadable or open in Excel — assume it matters rather than clobber it.
        return True
    for name in wb.sheetnames:
        ws = wb[name]
        col = 2 if name == "Questions" else 1
        for row in range(5, ws.max_row + 1):
            if name != "Questions" and ws.cell(row=row, column=2).value in (
                None,
                "",
                "Metric",
            ):
                continue
            if ws.cell(row=row, column=col).value not in (None, "", "Your value"):
                return True
    return False


def safe_out_path(payload: dict) -> Path:
    """Never overwrite a workbook that already has answers typed into it."""
    out = EXPORTS / OUT_NAME
    if not has_typed_answers(out):
        return out
    stamp = (payload.get("reportDate") or "new").replace("-", "")
    alt = EXPORTS / f"{OUT_NAME[:-5]} {stamp} rebuilt.xlsx"
    n = 2
    while alt.is_file():
        alt = EXPORTS / f"{OUT_NAME[:-5]} {stamp} rebuilt {n}.xlsx"
        n += 1
    print(
        f"Kept your filled copy: {out.name}\n"
        f"  writing the rebuild alongside it as {alt.name}"
    )
    return alt


def main() -> None:
    export = newest_export()
    payload = json.loads(export.read_text(encoding="utf-8"))
    as_of = payload.get("reportDate") or export.name[:10]

    wb = Workbook()
    wb.remove(wb.active)
    build_questions_sheet(wb)

    written = 0
    for agent in payload.get("agents", []):
        goals = agent.get("goals")
        if not goals:
            continue
        build_am_sheet(wb, agent["agentName"], goals, as_of)
        written += 1
    if not written:
        raise SystemExit(f"No Goals blocks found in {export.name}.")

    out = safe_out_path(payload)
    wb.save(out)
    print(f"Wrote {out}  ({written} AM sheets, source {export.name})")
    mirror_to_cursor("am_brief", out)


if __name__ == "__main__":
    main()
